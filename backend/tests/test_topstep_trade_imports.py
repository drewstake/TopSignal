from __future__ import annotations

import base64
import csv
import hashlib
import io
import logging
import os
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from threading import Barrier, Event
from pathlib import Path
from xml.sax.saxutils import escape
from time import perf_counter, sleep
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
import app.services.trade_imports as trade_imports_module
from sqlalchemy import create_engine, event
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")

from app.db import Base
from app.models import Account, ProjectXTradeEvent, TradeImportBatch, TradeImportPreview
from app.services.projectx_trades import (
    get_trade_event_pnl_calendar,
    list_trade_events,
    store_trade_events,
    summarize_trade_events,
)
from app.services.trade_imports import (
    TradeImportValidationError,
    cleanup_trade_import_previews,
    confirm_trade_import,
    get_trade_import_status,
    preview_trade_import,
)


USER_ID = "11111111-1111-1111-1111-111111111111"
OTHER_USER_ID = "22222222-2222-2222-2222-222222222222"
ACCOUNT_ID = 7301
MAX_5000_ROW_IMPORT_SECONDS = 15.0

TOPSTEP_COLUMNS = [
    "Id",
    "ContractName",
    "EnteredAt",
    "ExitedAt",
    "EntryPrice",
    "ExitPrice",
    "Fees",
    "PnL",
    "Size",
    "Type",
    "TradeDay",
    "TradeDuration",
    "Commissions",
]


@pytest.fixture()
def db_session():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    with engine.connect() as connection:
        connection.exec_driver_sql("PRAGMA foreign_keys=ON")
    Base.metadata.create_all(
        bind=engine,
        tables=[
            Account.__table__,
            TradeImportBatch.__table__,
            TradeImportPreview.__table__,
            ProjectXTradeEvent.__table__,
        ],
    )
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    session = SessionLocal()
    session.add(
        Account(
            id=ACCOUNT_ID,
            user_id=USER_ID,
            provider="projectx",
            external_id=str(ACCOUNT_ID),
            name="Topstep Live",
            trade_data_source="csv_import",
            account_state="ACTIVE",
        )
    )
    session.commit()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(
            bind=engine,
            tables=[
                ProjectXTradeEvent.__table__,
                TradeImportPreview.__table__,
                TradeImportBatch.__table__,
                Account.__table__,
            ],
        )
        engine.dispose()


def _trade_row(**overrides: str) -> dict[str, str]:
    row = {
        "Id": "2815118967",
        "ContractName": "MNQU6",
        "EnteredAt": "07/02/2026 10:10:08 -04:00",
        "ExitedAt": "07/02/2026 10:10:48 -04:00",
        "EntryPrice": "30182.500000000",
        "ExitPrice": "30148.750000000",
        "Fees": "2.22000",
        "PnL": "202.500000000",
        "Size": "3",
        "Type": "Short",
        "TradeDay": "07/02/2026 00:00:00 -05:00",
        "TradeDuration": "00:00:39.6715820",
        "Commissions": "1.50000",
    }
    row.update(overrides)
    return row


def _csv_bytes(
    rows: list[dict[str, str]],
    *,
    columns: list[tuple[str, str]] | None = None,
    bom: bool = True,
) -> bytes:
    mapped_columns = columns or [(column, column) for column in TOPSTEP_COLUMNS]
    output = io.StringIO(newline="")
    writer = csv.DictWriter(
        output,
        fieldnames=[output_name for output_name, _ in mapped_columns],
        lineterminator="\r\n",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                output_name: row[canonical_name]
                for output_name, canonical_name in mapped_columns
            }
        )
    encoding = "utf-8-sig" if bom else "utf-8"
    return output.getvalue().encode(encoding)


def _xlsx_bytes(rows: list[dict[str, str]]) -> bytes:
    """Build a minimal standards-compliant XLSX without a test-only dependency."""

    values = [TOPSTEP_COLUMNS, *[[row[column] for column in TOPSTEP_COLUMNS] for row in rows]]
    sheet_rows: list[str] = []
    for row_index, row in enumerate(values, start=1):
        cells = []
        for column_index, value in enumerate(row, start=1):
            reference = f"{_excel_column(column_index)}{row_index}"
            cells.append(
                f'<c r="{reference}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>'
            )
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData>'
        "</worksheet>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Trades" sheetId="1" r:id="rId1"/></sheets>'
        "</workbook>"
    )
    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" '
        'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )

    output = io.BytesIO()
    with ZipFile(output, mode="w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", root_rels)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return output.getvalue()


def _xlsx_with_declared_dimension(content: bytes, reference: str) -> bytes:
    source = io.BytesIO(content)
    output = io.BytesIO()
    with ZipFile(source, mode="r") as input_archive, ZipFile(
        output,
        mode="w",
        compression=ZIP_DEFLATED,
    ) as output_archive:
        for entry in input_archive.infolist():
            entry_content = input_archive.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                entry_content = entry_content.replace(
                    b"<sheetData>",
                    f'<dimension ref="{reference}"/><sheetData>'.encode("ascii"),
                    1,
                )
            output_archive.writestr(entry, entry_content)
    return output.getvalue()


def _typed_datetime_xlsx_bytes(
    *,
    entered_at: date | datetime,
    exited_at: date | datetime,
    entered_format: str,
    exited_format: str,
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Trades"
    worksheet.append(TOPSTEP_COLUMNS)
    row = _trade_row()
    worksheet.append([row[column] for column in TOPSTEP_COLUMNS])
    entered_cell = worksheet.cell(row=2, column=TOPSTEP_COLUMNS.index("EnteredAt") + 1)
    entered_cell.value = entered_at
    entered_cell.number_format = entered_format
    exited_cell = worksheet.cell(row=2, column=TOPSTEP_COLUMNS.index("ExitedAt") + 1)
    exited_cell.value = exited_at
    exited_cell.number_format = exited_format
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _excel_column(index: int) -> str:
    output = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        output = chr(65 + remainder) + output
    return output


def _preview(db_session, content: bytes, *, filename: str = "trades_export.csv"):
    return preview_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        filename=filename,
        content=content,
    )


def _confirm(db_session, content: bytes, *, filename: str = "trades_export.csv"):
    preview = _preview(db_session, content, filename=filename)
    result = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    return preview, result


def _iso_datetime(value: str | datetime) -> str:
    return value.isoformat() if isinstance(value, datetime) else value


def _iso_date(value: str | date) -> str:
    return value.isoformat() if isinstance(value, date) else value


def test_successful_csv_preview_and_confirm_persist_trade_and_import_metadata(db_session):
    content = _csv_bytes([_trade_row()])

    preview, confirmed = _confirm(db_session, content)

    assert preview["source_file_name"] == "trades_export.csv"
    assert preview["file_sha256"] == hashlib.sha256(content).hexdigest()
    assert preview["total_rows"] == 1
    assert preview["new_rows"] == 1
    assert preview["duplicate_rows"] == 0
    assert preview["summary"] == {
        "gross_pnl": 202.5,
        "fees": 2.22,
        "commissions": 1.5,
        "net_pnl": 198.78,
        "wins": 1,
        "losses": 0,
        "breakeven": 0,
    }

    trade = preview["trades"][0]
    assert trade["row_number"] == 2
    assert trade["source_trade_id"] == "2815118967"
    assert trade["contract_name"] == "MNQU6"
    assert trade["symbol"] == "MNQ"
    assert _iso_datetime(trade["entered_at"]) == "2026-07-02T14:10:08+00:00"
    assert _iso_datetime(trade["exited_at"]) == "2026-07-02T14:10:48+00:00"
    assert trade["entry_price"] == 30182.5
    assert trade["exit_price"] == 30148.75
    assert trade["gross_pnl"] == 202.5
    assert trade["net_pnl"] == 198.78
    assert trade["direction"].upper() == "SHORT"
    assert _iso_date(trade["trade_day"]) == "2026-07-02"
    assert trade["status"] == "new"

    assert confirmed["source_file_name"] == "trades_export.csv"
    assert confirmed["total_rows"] == 1
    assert confirmed["inserted_rows"] == 1
    assert confirmed["duplicate_rows"] == 0
    assert confirmed["import_id"] > 0
    imported_at = confirmed["imported_at"]
    if isinstance(imported_at, str):
        imported_at = datetime.fromisoformat(imported_at)
    assert imported_at.tzinfo is not None

    batch = db_session.query(TradeImportBatch).one()
    assert batch.id == confirmed["import_id"]
    assert batch.user_id == USER_ID
    assert batch.account_id == ACCOUNT_ID
    assert batch.source_file_name == "trades_export.csv"
    assert batch.file_sha256 == preview["file_sha256"]
    assert batch.imported_at is not None

    stored = db_session.query(ProjectXTradeEvent).one()
    assert stored.user_id == USER_ID
    assert stored.account_id == ACCOUNT_ID
    assert stored.source_trade_id == "2815118967"
    assert stored.order_id == "2815118967"
    assert stored.contract_id == "MNQU6"
    assert stored.symbol == "MNQ"
    # ProjectX events store the closing execution side.
    assert stored.side == "BUY"
    assert float(stored.size) == 3.0
    assert float(stored.entry_price) == 30182.5
    assert float(stored.price) == 30148.75
    assert float(stored.pnl) == 202.5
    assert float(stored.fees) == 2.22
    assert float(stored.commissions) == 1.5
    assert stored.fee_scope == "round_turn"
    assert stored.trade_date == date(2026, 7, 2)
    assert stored.import_batch_id == batch.id

    summary = summarize_trade_events(
        db_session,
        account_id=ACCOUNT_ID,
        user_id=USER_ID,
    )
    assert summary["avg_win_duration_minutes"] == 0.67


def test_same_file_and_overlapping_files_do_not_duplicate_or_overwrite_trades(db_session):
    first_row = _trade_row()
    overlapping_row = _trade_row(
        Id="2815266522",
        EnteredAt="07/02/2026 10:11:40 -04:00",
        ExitedAt="07/02/2026 10:18:09 -04:00",
        EntryPrice="30180.250000000",
        ExitPrice="30114.250000000",
        Fees="0.74000",
        PnL="132.000000000",
        Size="1",
        Commissions="0.50000",
    )
    third_row = _trade_row(
        Id="2815247404",
        EnteredAt="07/02/2026 10:12:00 -04:00",
        ExitedAt="07/02/2026 10:19:00 -04:00",
        EntryPrice="30180.250000000",
        ExitPrice="30147.000000000",
        Fees="2.96000",
        PnL="266.000000000",
        Size="4",
        Commissions="2.00000",
    )
    first_content = _csv_bytes([first_row, overlapping_row])
    first_preview, first_confirm = _confirm(db_session, first_content, filename="july-1.csv")

    repeated_preview = _preview(db_session, first_content, filename="renamed-copy.csv")
    repeated_confirm = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=repeated_preview["preview_token"],
    )

    assert first_preview["new_rows"] == 2
    assert repeated_preview["new_rows"] == 0
    assert repeated_preview["duplicate_rows"] == 2
    assert repeated_preview["summary"]["net_pnl"] == 0.0
    assert repeated_preview["summary"]["wins"] == 0
    assert [trade["status"] for trade in repeated_preview["trades"]] == [
        "duplicate",
        "duplicate",
    ]
    assert repeated_confirm["import_id"] == first_confirm["import_id"]
    assert db_session.query(ProjectXTradeEvent).count() == 2
    assert db_session.query(TradeImportBatch).count() == 1

    original_overlap = (
        db_session.query(ProjectXTradeEvent)
        .filter(ProjectXTradeEvent.source_trade_id == "2815266522")
        .one()
    )
    original_batch_id = original_overlap.import_batch_id
    original_created_at = original_overlap.created_at

    overlap_content = _csv_bytes([overlapping_row, third_row])
    overlap_preview, overlap_confirm = _confirm(
        db_session,
        overlap_content,
        filename="july-2.csv",
    )

    assert overlap_preview["new_rows"] == 1
    assert overlap_preview["duplicate_rows"] == 1
    assert overlap_preview["summary"] == {
        "gross_pnl": 266.0,
        "fees": 2.96,
        "commissions": 2.0,
        "net_pnl": 261.04,
        "wins": 1,
        "losses": 0,
        "breakeven": 0,
    }
    assert [trade["status"] for trade in overlap_preview["trades"]] == [
        "duplicate",
        "new",
    ]
    assert overlap_confirm["inserted_rows"] == 1
    assert overlap_confirm["duplicate_rows"] == 1
    assert db_session.query(ProjectXTradeEvent).count() == 3
    assert db_session.query(TradeImportBatch).count() == 2

    unchanged_overlap = (
        db_session.query(ProjectXTradeEvent)
        .filter(ProjectXTradeEvent.source_trade_id == "2815266522")
        .one()
    )
    assert unchanged_overlap.import_batch_id == original_batch_id
    assert unchanged_overlap.created_at == original_created_at
    assert float(unchanged_overlap.pnl) == 132.0


def test_live_daily_import_flow_remains_local_idempotent_and_conflict_safe(db_session):
    first_row = _trade_row()
    second_row = _trade_row(
        Id="2815118968",
        EnteredAt="07/02/2026 11:00:00 -04:00",
        ExitedAt="07/02/2026 11:01:00 -04:00",
        EntryPrice="30100.000000000",
        ExitPrice="30095.000000000",
        Fees="0.74000",
        PnL="-10.000000000",
        Size="1",
        Type="Long",
        TradeDuration="00:01:00.0000000",
        Commissions="0.50000",
    )
    third_row = _trade_row(
        Id="2815118969",
        EnteredAt="07/02/2026 12:00:00 -04:00",
        ExitedAt="07/02/2026 12:02:00 -04:00",
        EntryPrice="30100.000000000",
        ExitPrice="30105.000000000",
        Fees="0.74000",
        PnL="10.000000000",
        Size="1",
        Type="Long",
        TradeDuration="00:02:00.0000000",
        Commissions="0.50000",
    )

    daily_export = _csv_bytes([first_row, second_row])
    initial_preview, initial_result = _confirm(
        db_session,
        daily_export,
        filename="topstep-daily.csv",
    )
    assert initial_preview["new_rows"] == 2
    assert initial_result["inserted_rows"] == 2

    # These are the same local read services used by the dashboard/trades and
    # journal-facing calendar after an import; no provider client participates.
    local_summary = summarize_trade_events(
        db_session,
        account_id=ACCOUNT_ID,
        user_id=USER_ID,
    )
    local_trades = list_trade_events(
        db_session,
        account_id=ACCOUNT_ID,
        user_id=USER_ID,
        limit=200,
    )
    local_calendar = get_trade_event_pnl_calendar(
        db_session,
        account_id=ACCOUNT_ID,
        user_id=USER_ID,
    )
    assert local_summary["trade_count"] == 2
    assert len(local_trades) == 2
    assert local_calendar[0]["trade_count"] == 2

    # A renamed exact export gives the API shape used by the compact duplicate
    # state and resolves to the original durable result without another batch.
    duplicate_preview = _preview(
        db_session,
        daily_export,
        filename="renamed-topstep-daily.csv",
    )
    duplicate_result = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=duplicate_preview["preview_token"],
    )
    assert duplicate_preview["new_rows"] == 0
    assert duplicate_preview["duplicate_rows"] == 2
    assert duplicate_preview["conflict_rows"] == 0
    assert duplicate_result["import_id"] == initial_result["import_id"]

    overlap_preview, overlap_result = _confirm(
        db_session,
        _csv_bytes([second_row, third_row]),
        filename="overlap.csv",
    )
    assert [row["status"] for row in overlap_preview["trades"]] == ["duplicate", "new"]
    assert overlap_result["inserted_rows"] == 1
    assert overlap_result["duplicate_rows"] == 1

    conflict_preview = _preview(
        db_session,
        _csv_bytes([_trade_row(PnL="999.000000000")]),
        filename="conflict.csv",
    )
    assert conflict_preview["new_rows"] == 0
    assert conflict_preview["conflict_rows"] == 1
    with pytest.raises(TradeImportValidationError) as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=conflict_preview["preview_token"],
        )
    assert exc_info.value.code == "import_conflicts_unresolved"
    assert db_session.query(ProjectXTradeEvent).count() == 3
    assert db_session.query(TradeImportBatch).count() == 2


def test_same_identity_with_different_economics_is_a_blocking_conflict(db_session):
    existing = ProjectXTradeEvent(
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        contract_id="CON.F.US.MNQ.U26",
        symbol="MNQ",
        side="SELL",
        size=1,
        price=30000,
        trade_timestamp=datetime(2026, 7, 2, 14, 10, 48, tzinfo=timezone.utc),
        fees=0.37,
        fee_scope="per_side",
        pnl=999,
        order_id="provider-order",
        source_trade_id="2815118967",
        raw_payload={"source": "projectx"},
    )
    db_session.add(existing)
    db_session.commit()
    existing_id = int(existing.id)
    existing_created_at = existing.created_at

    content = _csv_bytes([_trade_row()])
    preview = _preview(db_session, content)

    assert preview["new_rows"] == 0
    assert preview["duplicate_rows"] == 0
    assert preview["conflict_rows"] == 1
    assert preview["summary"]["net_pnl"] == 0.0
    assert preview["trades"][0]["status"] == "conflict"
    assert {item["field"] for item in preview["trades"][0]["conflict"]["differences"]} >= {
        "contract",
        "quantity",
        "exit_price",
        "gross_pnl",
    }
    with pytest.raises(TradeImportValidationError, match="conflicting trade identities") as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
    assert exc_info.value.code == "import_conflicts_unresolved"
    assert db_session.query(ProjectXTradeEvent).count() == 1
    assert db_session.query(TradeImportBatch).count() == 0

    unchanged = db_session.query(ProjectXTradeEvent).one()
    assert unchanged.id == existing_id
    assert unchanged.created_at == existing_created_at
    assert unchanged.order_id == "provider-order"
    assert unchanged.import_batch_id is None
    assert float(unchanged.pnl) == 999.0
    assert unchanged.raw_payload == {"source": "projectx"}


def test_equal_fallback_order_and_exit_identity_is_a_harmless_duplicate(db_session):
    db_session.add(
        ProjectXTradeEvent(
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            contract_id="MNQU6",
            symbol="MNQ",
            side="BUY",
            size=3,
            price=30148.75,
            trade_timestamp=datetime(2026, 7, 2, 14, 10, 48, tzinfo=timezone.utc),
            entry_timestamp=datetime(2026, 7, 2, 14, 10, 8, tzinfo=timezone.utc),
            entry_price=30182.5,
            fees=2.22,
            commissions=1.5,
            fee_scope="round_turn",
            pnl=202.5,
            trade_date=date(2026, 7, 2),
            order_id="2815118967",
            source_trade_id="provider-source-id",
            raw_payload={"source": "projectx"},
        )
    )
    db_session.commit()
    content = _csv_bytes([_trade_row()])

    preview, confirmed = _confirm(db_session, content)

    assert preview["new_rows"] == 0
    assert preview["duplicate_rows"] == 1
    assert preview["trades"][0]["status"] == "duplicate"
    assert confirmed["inserted_rows"] == 0
    assert confirmed["duplicate_rows"] == 1
    assert db_session.query(ProjectXTradeEvent).count() == 1


def test_provider_sync_cannot_overwrite_a_confirmed_import(db_session):
    content = _csv_bytes([_trade_row()])
    _, confirmed = _confirm(db_session, content)
    original = db_session.query(ProjectXTradeEvent).one()
    original_payload = dict(original.raw_payload)

    inserted = store_trade_events(
        db_session,
        [
            {
                "account_id": ACCOUNT_ID,
                "contract_id": "CHANGED",
                "symbol": "ES",
                "side": "SELL",
                "size": 99,
                "price": 1,
                "timestamp": datetime(2026, 7, 2, 14, 10, 48, tzinfo=timezone.utc),
                "fees": 99,
                "pnl": -999,
                "order_id": "provider-order",
                "source_trade_id": "2815118967",
                "raw_payload": {"source": "projectx"},
            }
        ],
        user_id=USER_ID,
    )
    db_session.commit()

    stored = db_session.query(ProjectXTradeEvent).one()
    assert inserted == 0
    assert stored.import_batch_id == confirmed["import_id"]
    assert stored.contract_id == "MNQU6"
    assert stored.order_id == "2815118967"
    assert float(stored.pnl) == 202.5
    assert float(stored.fees) == 2.22
    assert float(stored.commissions) == 1.5
    assert stored.trade_date == date(2026, 7, 2)
    assert stored.raw_payload == original_payload


def test_fee_net_and_multiple_trades_same_day_flow_into_pnl_calendar(db_session):
    loss = _trade_row(
        Id="2815118968",
        EnteredAt="07/02/2026 11:00:00 -04:00",
        ExitedAt="07/02/2026 11:01:00 -04:00",
        EntryPrice="30100.000000000",
        ExitPrice="30095.000000000",
        Fees="0.74000",
        PnL="-10.000000000",
        Size="1",
        Type="Long",
        TradeDuration="00:01:00.0000000",
        Commissions="0.50000",
    )
    content = _csv_bytes([_trade_row(), loss])

    preview, _ = _confirm(db_session, content)

    assert preview["summary"] == {
        "gross_pnl": 192.5,
        "fees": 2.96,
        "commissions": 2.0,
        "net_pnl": 187.54,
        "wins": 1,
        "losses": 1,
        "breakeven": 0,
    }
    assert db_session.query(ProjectXTradeEvent).count() == 2

    calendar = get_trade_event_pnl_calendar(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
    )
    assert calendar == [
        {
            "date": "2026-07-02",
            "trade_count": 2,
            "win_count": 1,
            "loss_count": 1,
            "breakeven_count": 0,
            "gross_pnl": 192.5,
            "non_commission_fees": 2.96,
            "commissions": 2.0,
            "fees": 4.96,
            "net_pnl": 187.54,
        }
    ]


def test_identical_economic_values_with_distinct_topstep_ids_are_both_imported(db_session):
    # The real Topstep sample contains this shape: every economic field can be
    # identical even though each row represents a legitimate, distinct trade.
    first = _trade_row(Id="2815502338")
    second = _trade_row(Id="2815502261")
    content = _csv_bytes([first, second])

    preview, confirmed = _confirm(db_session, content)

    assert preview["new_rows"] == 2
    assert confirmed["inserted_rows"] == 2
    assert {
        row.source_trade_id
        for row in db_session.query(ProjectXTradeEvent).all()
    } == {"2815502338", "2815502261"}


def test_duplicate_identity_is_scoped_to_authenticated_user_and_account(db_session):
    db_session.add(
        ProjectXTradeEvent(
            user_id=OTHER_USER_ID,
            account_id=ACCOUNT_ID,
            contract_id="MNQU6",
            symbol="MNQ",
            side="BUY",
            size=3,
            price=30148.75,
            trade_timestamp=datetime(2026, 7, 2, 14, 10, 48, tzinfo=timezone.utc),
            fees=2.22,
            commissions=1.5,
            fee_scope="round_turn",
            pnl=202.5,
            order_id="2815118967",
            source_trade_id="2815118967",
        )
    )
    db_session.commit()

    content = _csv_bytes([_trade_row()])
    preview, confirmed = _confirm(db_session, content)

    assert preview["new_rows"] == 1
    assert preview["duplicate_rows"] == 0
    assert confirmed["inserted_rows"] == 1
    assert db_session.query(ProjectXTradeEvent).count() == 2


def test_parser_accepts_reordered_friendly_column_aliases(db_session):
    aliases = [
        ("Commission", "Commissions"),
        ("Trade Duration", "TradeDuration"),
        ("Trade Day", "TradeDay"),
        ("Direction", "Type"),
        ("Quantity", "Size"),
        ("P&L", "PnL"),
        ("Fee", "Fees"),
        ("Exit Price", "ExitPrice"),
        ("Entry Price", "EntryPrice"),
        ("Exited At", "ExitedAt"),
        ("Entered At", "EnteredAt"),
        ("Contract Name", "ContractName"),
        ("Trade Id", "Id"),
    ]
    content = _csv_bytes([_trade_row()], columns=aliases, bom=False)

    preview = _preview(db_session, content, filename="friendly-columns.CSV")

    assert preview["total_rows"] == 1
    assert preview["new_rows"] == 1
    assert preview["trades"][0]["source_trade_id"] == "2815118967"
    assert preview["trades"][0]["net_pnl"] == 198.78


def test_parser_accepts_excel_open_xml_workbook(db_session):
    filename = "trades_export.XLSX"
    content = _xlsx_bytes([_trade_row()])

    preview = _preview(db_session, content, filename=filename)

    assert preview["source_file_name"] == filename
    assert preview["file_sha256"] == hashlib.sha256(content).hexdigest()
    assert preview["total_rows"] == 1
    assert preview["new_rows"] == 1
    assert preview["summary"]["net_pnl"] == 198.78


@pytest.mark.parametrize("dimension", ["A1:XFD1048576", "A1:A1048576", "A1:XFD2"])
def test_parser_rejects_malicious_xlsx_dimensions_before_iteration(
    db_session,
    dimension,
):
    content = _xlsx_with_declared_dimension(
        _xlsx_bytes([_trade_row()]),
        dimension,
    )

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content, filename="oversized-dimension.xlsx")

    assert exc_info.value.code == "excel_worksheet_too_large"
    assert db_session.query(ProjectXTradeEvent).count() == 0


def test_parser_rejects_unverified_macro_enabled_workbooks(db_session):
    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(
            db_session,
            _xlsx_bytes([_trade_row()]),
            filename="trades_export.xlsm",
        )

    assert exc_info.value.code == "unsupported_file_type"
    assert db_session.query(ProjectXTradeEvent).count() == 0


@pytest.mark.parametrize("date_only_field", ["EnteredAt", "ExitedAt"])
def test_parser_rejects_date_only_formatted_xlsx_trade_timestamps(
    db_session,
    date_only_field,
):
    entered_at: date | datetime = datetime(2026, 7, 2, 10, 10, 8)
    exited_at: date | datetime = datetime(2026, 7, 2, 10, 10, 48)
    entered_format = "mm/dd/yyyy hh:mm:ss"
    exited_format = "mm/dd/yyyy hh:mm:ss"
    if date_only_field == "EnteredAt":
        entered_at = date(2026, 7, 2)
        entered_format = "mm/dd/yyyy"
    else:
        exited_at = date(2026, 7, 2)
        exited_format = "mm/dd/yyyy"

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(
            db_session,
            _typed_datetime_xlsx_bytes(
                entered_at=entered_at,
                exited_at=exited_at,
                entered_format=entered_format,
                exited_format=exited_format,
            ),
            filename="date-only.xlsx",
        )

    assert exc_info.value.code == "invalid_rows"
    assert exc_info.value.row_errors[0]["field"] == date_only_field
    assert "date-only spreadsheet" in exc_info.value.row_errors[0]["message"].lower()


def test_parser_accepts_genuine_midnight_xlsx_timestamps_with_time_format(db_session):
    preview = _preview(
        db_session,
        _typed_datetime_xlsx_bytes(
            entered_at=datetime(2026, 7, 2, 0, 0, 0),
            exited_at=datetime(2026, 7, 2, 0, 1, 0),
            entered_format="mm/dd/yyyy hh:mm:ss",
            exited_format="mm/dd/yyyy hh:mm:ss",
        ),
        filename="midnight.xlsx",
    )

    assert _iso_datetime(preview["trades"][0]["entered_at"]) == "2026-07-02T04:00:00+00:00"
    assert _iso_datetime(preview["trades"][0]["exited_at"]) == "2026-07-02T04:01:00+00:00"


@pytest.mark.parametrize(
    "fixture_name",
    [
        "topstep_trade_export_utf8.csv",
        "topstep_trade_export_utf8_bom.csv",
        "topstep_trade_export_semicolon.csv",
        "topstep_trade_export_tab.csv",
        "topstep_trade_export_pipe.csv",
    ],
)
def test_golden_csv_topstep_fixtures_are_supported(db_session, fixture_name):
    fixture = Path(__file__).parent / "fixtures" / fixture_name
    preview = _preview(db_session, fixture.read_bytes(), filename=fixture.name)

    assert preview["total_rows"] == 1
    assert preview["new_rows"] == 1
    assert preview["trades"][0]["source_trade_id"] == "2815118967"
    assert preview["summary"]["net_pnl"] == 198.78


def test_golden_xlsx_topstep_fixture_is_supported(db_session):
    fixture = Path(__file__).parent / "fixtures" / "topstep_trade_export_xlsx.base64"
    content = base64.b64decode(fixture.read_text(encoding="ascii").strip(), validate=True)
    upload_name = "topstep_trade_export.xlsx"

    preview = _preview(db_session, content, filename=upload_name)

    assert preview["source_file_name"] == upload_name
    assert preview["total_rows"] == 1
    assert preview["new_rows"] == 1
    assert preview["trades"][0]["source_trade_id"] == "2815118967"
    assert preview["summary"]["net_pnl"] == 198.78


@pytest.mark.parametrize("encoding", ["utf-16", "cp1252"])
def test_unverified_csv_encodings_are_rejected_strictly(db_session, encoding):
    text = _csv_bytes([_trade_row()]).decode("utf-8-sig") + "\n# café"
    content = text.encode(encoding)

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content, filename="unsupported-encoding.csv")

    assert exc_info.value.code == "invalid_csv_encoding"
    assert db_session.query(ProjectXTradeEvent).count() == 0


@pytest.mark.parametrize(
    ("filename", "content"),
    [
        ("trades.txt", b"not a supported trade export"),
        ("trades.xlsx", b"not an xlsx zip archive"),
        ("trades.csv", b"\xff\xfe\x00\x01"),
        ("trades.csv", b""),
    ],
)
def test_invalid_files_are_rejected_without_database_writes(
    db_session,
    filename,
    content,
):
    with pytest.raises(TradeImportValidationError):
        _preview(db_session, content, filename=filename)

    assert db_session.query(ProjectXTradeEvent).count() == 0
    assert db_session.query(TradeImportBatch).count() == 0


def test_missing_required_columns_names_them_clearly(db_session):
    columns = [
        (column, column)
        for column in TOPSTEP_COLUMNS
        if column not in {"Id", "PnL"}
    ]
    content = _csv_bytes([_trade_row()], columns=columns)

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content)

    message = str(exc_info.value).lower()
    assert "missing" in message
    assert "id" in message
    assert "pnl" in message or "p&l" in message
    assert db_session.query(ProjectXTradeEvent).count() == 0
    assert db_session.query(TradeImportBatch).count() == 0


def test_invalid_row_reports_row_number_and_field(db_session):
    content = _csv_bytes([_trade_row(PnL="not-a-number")])

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content)

    message = str(exc_info.value).lower()
    assert "row 2" in message
    assert "pnl" in message or "p&l" in message
    assert db_session.query(ProjectXTradeEvent).count() == 0


def test_oversized_or_control_character_trade_ids_are_rejected(db_session):
    for source_trade_id in ("x" * 256, "unsafe\x00id"):
        content = _csv_bytes([_trade_row(Id=source_trade_id)])

        with pytest.raises(TradeImportValidationError) as exc_info:
            _preview(db_session, content)

        assert exc_info.value.code == "invalid_rows"
        assert exc_info.value.row_errors[0]["field"] == "Id"


def test_nul_in_an_extra_raw_column_is_rejected_before_database_insert(db_session):
    content = _csv_bytes(
        [_trade_row(Notes="unsafe\x00note")],
        columns=[
            *((column, column) for column in TOPSTEP_COLUMNS),
            ("Notes", "Notes"),
        ],
    )

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content)

    assert exc_info.value.code == "invalid_rows"
    assert exc_info.value.row_errors[0]["field"] == "Notes"


@pytest.mark.parametrize("field", ["EntryPrice", "PnL", "Fees", "Commissions", "Size"])
def test_numeric_values_outside_database_range_are_rejected(db_session, field):
    content = _csv_bytes([_trade_row(**{field: "1e10000"})])

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content)

    assert exc_info.value.code == "invalid_rows"
    assert exc_info.value.row_errors[0]["field"] == field
    assert "supported numeric range" in exc_info.value.row_errors[0]["message"]


def test_confirm_uses_the_staged_manifest_without_a_second_upload(db_session):
    reviewed = _preview(db_session, _csv_bytes([_trade_row()]))

    confirmed = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=reviewed["preview_token"],
    )
    assert confirmed["inserted_rows"] == 1
    assert db_session.query(ProjectXTradeEvent).count() == 1
    assert db_session.query(TradeImportBatch).count() == 1


@pytest.mark.parametrize(
    ("override", "expected_field"),
    [
        ({"PnL": "203.50"}, "gross_pnl"),
        ({"Fees": "3.22"}, "fees"),
        ({"Commissions": "2.50"}, "commissions"),
        ({"Size": "4"}, "quantity"),
        ({"EnteredAt": "07/02/2026 10:09:08 -04:00"}, "entered_at"),
        ({"ExitedAt": "07/02/2026 10:11:48 -04:00"}, "exited_at"),
        ({"EntryPrice": "30183.50"}, "entry_price"),
        ({"ExitPrice": "30149.75"}, "exit_price"),
        ({"Type": "Long"}, "direction"),
        ({"ContractName": "MESU6"}, "contract"),
    ],
)
def test_corrected_same_id_economic_fields_are_conflicts(
    db_session,
    override,
    expected_field,
):
    _confirm(db_session, _csv_bytes([_trade_row()]))

    preview = _preview(db_session, _csv_bytes([_trade_row(**override)]), filename="correction.csv")

    assert preview["conflict_rows"] == 1
    assert preview["new_rows"] == 0
    assert preview["duplicate_rows"] == 0
    conflict = preview["trades"][0]["conflict"]
    assert conflict["reason"] == "stored_trade_mismatch"
    assert expected_field in {difference["field"] for difference in conflict["differences"]}
    with pytest.raises(TradeImportValidationError) as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
    assert exc_info.value.code == "import_conflicts_unresolved"
    assert db_session.query(ProjectXTradeEvent).count() == 1


def test_stored_trade_day_difference_is_a_conflict(db_session):
    _confirm(db_session, _csv_bytes([_trade_row()]))
    stored = db_session.query(ProjectXTradeEvent).one()
    stored.trade_date = date(2026, 7, 1)
    db_session.commit()

    preview = _preview(db_session, _csv_bytes([_trade_row()]), filename="trade-day-correction.csv")

    assert preview["conflict_rows"] == 1
    assert "trade_day" in {
        difference["field"]
        for difference in preview["trades"][0]["conflict"]["differences"]
    }


def test_repeated_id_inside_one_file_is_exact_duplicate_or_blocking_conflict(db_session):
    exact = _preview(db_session, _csv_bytes([_trade_row(), _trade_row()]), filename="exact.csv")
    assert [row["status"] for row in exact["trades"]] == ["new", "duplicate"]
    exact_result = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=exact["preview_token"],
    )
    assert exact_result["inserted_rows"] == 1
    assert exact_result["duplicate_rows"] == 1

    second_db = db_session.bind
    db_session.query(TradeImportPreview).delete()
    db_session.query(ProjectXTradeEvent).delete()
    db_session.query(TradeImportBatch).delete()
    db_session.commit()
    assert second_db is not None

    split_fill_shape = _preview(
        db_session,
        _csv_bytes([_trade_row(), _trade_row(Size="1", PnL="67.50")]),
        filename="unproven-split-fill.csv",
    )
    assert [row["status"] for row in split_fill_shape["trades"]] == ["new", "conflict"]
    assert split_fill_shape["trades"][1]["conflict"]["reason"] == "repeated_id_mismatch"
    with pytest.raises(TradeImportValidationError) as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=split_fill_shape["preview_token"],
        )
    assert exc_info.value.code == "import_conflicts_unresolved"
    assert db_session.query(ProjectXTradeEvent).count() == 0


def test_preview_becomes_stale_when_an_overlapping_import_changes_dedupe_state(db_session):
    content = _csv_bytes([_trade_row()])
    first_preview = _preview(db_session, content, filename="first.csv")
    _, concurrent_result = _confirm(db_session, content, filename="concurrent.csv")
    assert concurrent_result["inserted_rows"] == 1

    with pytest.raises(TradeImportValidationError) as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=first_preview["preview_token"],
        )

    assert exc_info.value.code == "preview_stale"
    status = get_trade_import_status(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=first_preview["preview_token"],
    )
    assert status["status"] == "stale"
    assert db_session.query(ProjectXTradeEvent).count() == 1


def test_status_recovers_committed_outcome_and_same_token_retry_is_idempotent(db_session):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    pending = get_trade_import_status(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    assert pending["status"] == "pending"
    assert pending["confirmation_retryable"] is True
    assert pending["result"] is None

    first = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    retry = confirm_trade_import(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    recovered = get_trade_import_status(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )

    assert retry == first
    assert recovered["status"] == "committed"
    assert recovered["confirmation_retryable"] is False
    assert recovered["result"] == first
    staged = db_session.query(TradeImportPreview).one()
    assert staged.normalized_manifest is None
    assert staged.preview_rows is None
    assert staged.dedupe_snapshot is None
    assert staged.token_hash != preview["preview_token"]
    assert db_session.query(ProjectXTradeEvent).count() == 1
    assert db_session.query(TradeImportBatch).count() == 1


def test_simultaneous_confirmation_inserts_once_and_recovers_one_outcome(tmp_path):
    database_path = tmp_path / "simultaneous-import.sqlite3"
    engine = create_engine(
        f"sqlite+pysqlite:///{database_path}",
        connect_args={"check_same_thread": False, "timeout": 15},
    )
    with engine.connect() as connection:
        connection.exec_driver_sql("PRAGMA foreign_keys=ON")
        connection.exec_driver_sql("PRAGMA journal_mode=WAL")
    Base.metadata.create_all(
        bind=engine,
        tables=[
            Account.__table__,
            TradeImportBatch.__table__,
            TradeImportPreview.__table__,
            ProjectXTradeEvent.__table__,
        ],
    )
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    setup = SessionLocal()
    try:
        setup.add(
            Account(
                id=9001,
                user_id=USER_ID,
                provider="projectx",
                external_id=str(ACCOUNT_ID),
                name="Topstep Live",
                trade_data_source="csv_import",
                account_state="ACTIVE",
            )
        )
        setup.commit()
        preview = preview_trade_import(
            setup,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            filename="simultaneous.csv",
            content=_csv_bytes([_trade_row()]),
        )
    finally:
        setup.close()

    barrier = Barrier(2)

    def confirm_once():
        session = SessionLocal()
        try:
            session.connection().exec_driver_sql("PRAGMA foreign_keys=ON")
            barrier.wait(timeout=10)
            try:
                return confirm_trade_import(
                    session,
                    user_id=USER_ID,
                    account_id=ACCOUNT_ID,
                    preview_token=preview["preview_token"],
                )
            except TradeImportValidationError as exc:
                return exc.code
        finally:
            session.close()

    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            outcomes = list(executor.map(lambda _index: confirm_once(), range(2)))
        results = [outcome for outcome in outcomes if isinstance(outcome, dict)]
        assert results
        assert all(result["import_id"] == results[0]["import_id"] for result in results)
        assert all(
            outcome == "confirmation_in_progress" or isinstance(outcome, dict)
            for outcome in outcomes
        )
        verify = SessionLocal()
        try:
            assert verify.query(ProjectXTradeEvent).count() == 1
            assert verify.query(TradeImportBatch).count() == 1
            status = get_trade_import_status(
                verify,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )
            assert status["status"] == "committed"
        finally:
            verify.close()
    finally:
        engine.dispose()


@pytest.mark.parametrize("contender_kind", ["cleanup", "status", "confirm"])
def test_expiration_cannot_overwrite_a_confirmation_that_commits_while_update_waits(
    tmp_path,
    monkeypatch,
    contender_kind,
):
    database_path = tmp_path / f"{contender_kind}-confirm-race.sqlite3"
    engine = create_engine(
        f"sqlite+pysqlite:///{database_path}",
        connect_args={"check_same_thread": False, "timeout": 15},
    )
    with engine.connect() as connection:
        connection.exec_driver_sql("PRAGMA foreign_keys=ON")
        connection.exec_driver_sql("PRAGMA journal_mode=WAL")
    Base.metadata.create_all(
        bind=engine,
        tables=[
            Account.__table__,
            TradeImportBatch.__table__,
            TradeImportPreview.__table__,
            ProjectXTradeEvent.__table__,
        ],
    )
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    setup = SessionLocal()
    try:
        setup.add(
            Account(
                id=9002,
                user_id=USER_ID,
                provider="projectx",
                external_id=str(ACCOUNT_ID),
                name="Topstep Live",
                trade_data_source="csv_import",
                account_state="ACTIVE",
            )
        )
        setup.commit()
        if contender_kind != "cleanup":
            monkeypatch.setattr(
                trade_imports_module,
                "TRADE_IMPORT_PREVIEW_TTL",
                timedelta(milliseconds=250),
            )
        preview = preview_trade_import(
            setup,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            filename="cleanup-race.csv",
            content=_csv_bytes([_trade_row()]),
        )
    finally:
        setup.close()

    confirmation_holds_claim = Event()
    release_confirmation = Event()
    expiration_update_started = Event()
    persist = trade_imports_module._persist_staged_import

    def pause_before_persist(*args, **kwargs):
        confirmation_holds_claim.set()
        assert release_confirmation.wait(timeout=10)
        return persist(*args, **kwargs)

    def observe_cleanup_update(
        _connection,
        _cursor,
        statement,
        _parameters,
        _context,
        _executemany,
    ):
        normalized = statement.lower()
        if "update trade_import_previews" in normalized and "expires_at" in normalized:
            expiration_update_started.set()

    event.listen(engine, "before_cursor_execute", observe_cleanup_update)
    monkeypatch.setattr(trade_imports_module, "_persist_staged_import", pause_before_persist)

    def confirm_once():
        session = SessionLocal()
        try:
            session.connection().exec_driver_sql("PRAGMA foreign_keys=ON")
            return confirm_trade_import(
                session,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )
        finally:
            session.close()

    def contend_once():
        session = SessionLocal()
        try:
            session.connection().exec_driver_sql("PRAGMA foreign_keys=ON")
            if contender_kind == "cleanup":
                return cleanup_trade_import_previews(
                    session,
                    now=datetime.now(timezone.utc) + timedelta(days=1),
                )
            if contender_kind == "status":
                return get_trade_import_status(
                    session,
                    user_id=USER_ID,
                    account_id=ACCOUNT_ID,
                    preview_token=preview["preview_token"],
                )
            return confirm_trade_import(
                session,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )
        finally:
            session.close()

    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            confirmation_future = executor.submit(confirm_once)
            assert confirmation_holds_claim.wait(timeout=10)
            if contender_kind != "cleanup":
                expires_at = preview["expires_at"]
                if expires_at.tzinfo is None:
                    expires_at = expires_at.replace(tzinfo=timezone.utc)
                while datetime.now(timezone.utc) <= expires_at:
                    sleep(0.01)
            contender_future = executor.submit(contend_once)
            assert expiration_update_started.wait(timeout=10)
            release_confirmation.set()
            confirmed = confirmation_future.result(timeout=15)
            contender_result = contender_future.result(timeout=15)

        assert confirmed["inserted_rows"] == 1
        if contender_kind == "cleanup":
            assert contender_result["expired_previews"] == 0
        elif contender_kind == "status":
            assert contender_result["status"] == "committed"
            assert contender_result["result"] == confirmed
        else:
            assert contender_result == confirmed
        verify = SessionLocal()
        try:
            status = get_trade_import_status(
                verify,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )
            assert status["status"] == "committed"
            assert verify.query(ProjectXTradeEvent).count() == 1
            assert verify.query(TradeImportBatch).count() == 1
        finally:
            verify.close()
    finally:
        release_confirmation.set()
        event.remove(engine, "before_cursor_execute", observe_cleanup_update)
        engine.dispose()


def test_status_is_restart_safe_and_token_is_account_scoped(db_session):
    preview, confirmed = _confirm(db_session, _csv_bytes([_trade_row()]))
    SessionLocal = sessionmaker(bind=db_session.bind, autoflush=False, autocommit=False)
    restarted = SessionLocal()
    try:
        status = get_trade_import_status(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert status["result"] == confirmed
    finally:
        restarted.close()

    db_session.add(
        Account(
            id=7302,
            user_id=USER_ID,
            provider="projectx",
            external_id="7302",
            name="Other Live",
            trade_data_source="csv_import",
            account_state="ACTIVE",
        )
    )
    db_session.commit()
    with pytest.raises(TradeImportValidationError) as exc_info:
        get_trade_import_status(
            db_session,
            user_id=USER_ID,
            account_id=7302,
            preview_token=preview["preview_token"],
        )
    assert exc_info.value.code == "preview_not_found"


def test_preview_survives_restart_before_confirmation(db_session):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    SessionLocal = sessionmaker(bind=db_session.bind, autoflush=False, autocommit=False)

    restarted = SessionLocal()
    try:
        pending = get_trade_import_status(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert pending["status"] == "pending"
        assert pending["confirmation_retryable"] is True

        confirmed = confirm_trade_import(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert confirmed["inserted_rows"] == 1
    finally:
        restarted.close()


def test_confirmation_rollback_is_retryable_after_restart(db_session, monkeypatch):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))

    def interrupt_before_commit(*_args, **_kwargs):
        raise RuntimeError("simulated interruption before commit")

    with monkeypatch.context() as patcher:
        patcher.setattr(trade_imports_module, "_persist_staged_import", interrupt_before_commit)
        with pytest.raises(RuntimeError, match="before commit"):
            confirm_trade_import(
                db_session,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )

    assert db_session.query(ProjectXTradeEvent).count() == 0
    assert db_session.query(TradeImportBatch).count() == 0
    assert db_session.query(TradeImportPreview).one().status == "pending"

    SessionLocal = sessionmaker(bind=db_session.bind, autoflush=False, autocommit=False)
    restarted = SessionLocal()
    try:
        status = get_trade_import_status(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert status["status"] == "pending"
        assert status["confirmation_retryable"] is True

        confirmed = confirm_trade_import(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        retry = confirm_trade_import(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert retry == confirmed
        assert restarted.query(ProjectXTradeEvent).count() == 1
        assert restarted.query(TradeImportBatch).count() == 1
    finally:
        restarted.close()


def test_committed_outcome_survives_lost_confirmation_response(db_session, monkeypatch):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    persist = trade_imports_module._persist_staged_import

    def commit_then_interrupt(*args, **kwargs):
        persist(*args, **kwargs)
        raise ConnectionError("simulated response interruption after commit")

    with monkeypatch.context() as patcher:
        patcher.setattr(trade_imports_module, "_persist_staged_import", commit_then_interrupt)
        with pytest.raises(ConnectionError, match="after commit"):
            confirm_trade_import(
                db_session,
                user_id=USER_ID,
                account_id=ACCOUNT_ID,
                preview_token=preview["preview_token"],
            )

    SessionLocal = sessionmaker(bind=db_session.bind, autoflush=False, autocommit=False)
    restarted = SessionLocal()
    try:
        recovered = get_trade_import_status(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert recovered["status"] == "committed"
        assert recovered["confirmation_retryable"] is False
        assert recovered["result"]["inserted_rows"] == 1

        retry = confirm_trade_import(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert retry == recovered["result"]
        assert restarted.query(ProjectXTradeEvent).count() == 1
        assert restarted.query(TradeImportBatch).count() == 1
    finally:
        restarted.close()


def test_restart_recovers_a_durable_legacy_confirming_marker(db_session):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    staged = db_session.query(TradeImportPreview).one()
    staged.status = "confirming"
    staged.outcome_code = None
    db_session.commit()

    SessionLocal = sessionmaker(bind=db_session.bind, autoflush=False, autocommit=False)
    restarted = SessionLocal()
    try:
        recovered = get_trade_import_status(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert recovered["status"] == "pending"
        assert recovered["outcome_code"] == "confirmation_retryable"
        assert recovered["confirmation_retryable"] is True

        confirmed = confirm_trade_import(
            restarted,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
        assert confirmed["inserted_rows"] == 1
    finally:
        restarted.close()


def test_expired_preview_is_scrubbed_and_cannot_confirm(db_session):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    staged = db_session.query(TradeImportPreview).one()
    staged.expires_at = datetime.now(timezone.utc) - timedelta(seconds=1)
    db_session.commit()

    status = get_trade_import_status(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    assert status["status"] == "expired"
    assert staged.normalized_manifest is None
    with pytest.raises(TradeImportValidationError) as exc_info:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
    assert exc_info.value.code == "preview_expired"


def test_scheduled_cleanup_scrubs_expired_manifests_and_enforces_retention(db_session):
    _preview(db_session, _csv_bytes([_trade_row()]), filename="expired.csv")
    _preview(
        db_session,
        _csv_bytes([_trade_row(Id="2815118968")]),
        filename="retained-too-long.csv",
    )
    first, second = db_session.query(TradeImportPreview).order_by(TradeImportPreview.id).all()
    now = datetime.now(timezone.utc)
    first.status = "confirming"
    first.expires_at = now - timedelta(minutes=1)
    first.retention_until = now + timedelta(days=1)
    second.expires_at = now - timedelta(days=8)
    second.retention_until = now - timedelta(seconds=1)
    db_session.commit()

    result = cleanup_trade_import_previews(db_session, now=now)

    assert result == {"expired_previews": 2, "deleted_previews": 1}
    remaining = db_session.query(TradeImportPreview).one()
    assert remaining.id == first.id
    assert remaining.status == "expired"
    assert remaining.outcome_code == "preview_expired"
    assert remaining.normalized_manifest is None
    assert remaining.preview_rows is None
    assert remaining.dedupe_snapshot is None


def test_archived_live_account_cannot_preview_or_confirm_but_status_remains_recoverable(db_session):
    preview = _preview(db_session, _csv_bytes([_trade_row()]))
    account = db_session.query(Account).filter(Account.external_id == str(ACCOUNT_ID)).one()
    account.archived_at = datetime.now(timezone.utc)
    db_session.commit()

    with pytest.raises(TradeImportValidationError) as preview_exc:
        _preview(db_session, _csv_bytes([_trade_row(Id="2815118968")]))
    assert preview_exc.value.code == "trade_import_account_archived"

    with pytest.raises(TradeImportValidationError) as confirm_exc:
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )
    assert confirm_exc.value.code == "trade_import_account_archived"

    status = get_trade_import_status(
        db_session,
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        preview_token=preview["preview_token"],
    )
    assert status["status"] == "pending"
    assert status["confirmation_retryable"] is False


@pytest.mark.parametrize(
    ("field", "value", "message_fragment"),
    [
        ("EnteredAt", "07/02/2026", "date-only"),
        ("ExitedAt", "2026-07-02", "date-only"),
        ("EnteredAt", "03/08/2026 02:30:00", "nonexistent"),
        ("ExitedAt", "11/01/2026 01:30:00", "ambiguous"),
    ],
)
def test_timestamp_validation_rejects_date_only_and_invalid_dst_wall_times(
    db_session,
    field,
    value,
    message_fragment,
):
    content = _csv_bytes([_trade_row(**{field: value})])

    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, content)

    assert exc_info.value.code == "invalid_rows"
    assert message_fragment in exc_info.value.row_errors[0]["message"].lower()


def test_unambiguous_naive_eastern_timestamps_remain_supported(db_session):
    preview = _preview(
        db_session,
        _csv_bytes(
            [
                _trade_row(
                    EnteredAt="07/02/2026 10:10:08",
                    ExitedAt="07/02/2026 10:10:48",
                    TradeDay="07/02/2026",
                )
            ]
        ),
    )
    assert _iso_datetime(preview["trades"][0]["entered_at"]) == "2026-07-02T14:10:08+00:00"


@pytest.mark.parametrize(
    ("entered_at", "exited_at", "trade_day"),
    [
        ("01/15/2026 10:00:00 -05:00", "01/15/2026 10:01:00 -05:00", "01/15/2026"),
        ("07/02/2026 10:00:00 -04:00", "07/02/2026 10:01:00 -04:00", "07/02/2026"),
        ("11/01/2026 01:10:00 -04:00", "11/01/2026 01:20:00 -04:00", "11/01/2026"),
        ("11/01/2026 01:10:00 -05:00", "11/01/2026 01:20:00 -05:00", "11/01/2026"),
        ("07/02/2026 17:58:00 -04:00", "07/02/2026 17:59:00 -04:00", "07/02/2026"),
        ("07/02/2026 18:00:00 -04:00", "07/02/2026 18:01:00 -04:00", "07/03/2026"),
        ("07/05/2026 18:00:00 -04:00", "07/05/2026 18:01:00 -04:00", "07/06/2026"),
        ("07/05/2026 20:00:00 -04:00", "07/06/2026 01:00:00 -04:00", "07/06/2026"),
    ],
)
def test_trade_day_is_validated_against_shared_six_pm_boundary(
    db_session,
    entered_at,
    exited_at,
    trade_day,
):
    preview = _preview(
        db_session,
        _csv_bytes(
            [
                _trade_row(
                    EnteredAt=entered_at,
                    ExitedAt=exited_at,
                    TradeDay=trade_day,
                )
            ]
        ),
    )
    assert _iso_date(preview["trades"][0]["trade_day"]) == datetime.strptime(
        trade_day,
        "%m/%d/%Y",
    ).date().isoformat()


def test_trade_day_mismatch_is_rejected(db_session):
    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(
            db_session,
            _csv_bytes(
                [
                    _trade_row(
                        EnteredAt="07/02/2026 18:00:00 -04:00",
                        ExitedAt="07/02/2026 18:01:00 -04:00",
                        TradeDay="07/02/2026",
                    )
                ]
            ),
        )
    assert exc_info.value.code == "invalid_rows"
    assert "6:00 pm et boundary" in exc_info.value.row_errors[0]["message"].lower()


@pytest.mark.parametrize("quantity", ["0", "-1", "1.5", "10001", "not-a-number"])
def test_quantity_must_be_a_positive_bounded_whole_contract_count(db_session, quantity):
    with pytest.raises(TradeImportValidationError) as exc_info:
        _preview(db_session, _csv_bytes([_trade_row(Size=quantity)]))
    assert exc_info.value.code == "invalid_rows"
    assert exc_info.value.row_errors[0]["field"] == "Size"


@pytest.mark.parametrize("quantity", ["1", "3.0", 10_000])
def test_integer_shaped_quantities_are_accepted(db_session, quantity):
    preview = _preview(db_session, _csv_bytes([_trade_row(Size=str(quantity))]))
    assert preview["trades"][0]["size"] == float(quantity)


def test_database_rejects_fractional_trade_event_quantity(db_session):
    db_session.add(
        ProjectXTradeEvent(
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            contract_id="MNQU6",
            symbol="MNQ",
            side="BUY",
            size=1.5,
            price=30148.75,
            trade_timestamp=datetime(2026, 7, 2, 14, 10, 48, tzinfo=timezone.utc),
            fees=0,
            order_id="fractional",
        )
    )
    with pytest.raises(IntegrityError):
        db_session.commit()
    db_session.rollback()


def test_database_rejects_orphan_batch_and_cross_account_event_links(db_session):
    orphan = TradeImportBatch(
        user_id=USER_ID,
        account_id=ACCOUNT_ID,
        account_row_id=999999,
        account_external_id=str(ACCOUNT_ID),
        source_file_name="orphan.csv",
        file_sha256="a" * 64,
        total_rows=0,
        inserted_rows=0,
        duplicate_rows=0,
    )
    db_session.add(orphan)
    with pytest.raises(IntegrityError):
        db_session.commit()
    db_session.rollback()

    _, confirmed = _confirm(db_session, _csv_bytes([_trade_row()]))
    other_user_account_row_id = 8301
    db_session.add(
        Account(
            id=other_user_account_row_id,
            user_id=OTHER_USER_ID,
            provider="projectx",
            external_id=str(ACCOUNT_ID),
            trade_data_source="csv_import",
            account_state="ACTIVE",
        )
    )
    db_session.commit()
    db_session.add(
        ProjectXTradeEvent(
            user_id=OTHER_USER_ID,
            account_id=ACCOUNT_ID,
            account_row_id=other_user_account_row_id,
            account_external_id=str(ACCOUNT_ID),
            contract_id="MNQU6",
            symbol="MNQ",
            side="BUY",
            size=1,
            price=1,
            trade_timestamp=datetime.now(timezone.utc),
            fees=0,
            order_id="cross-user",
            source_trade_id="cross-user",
            import_batch_id=confirmed["import_id"],
        )
    )
    with pytest.raises(IntegrityError):
        db_session.commit()
    db_session.rollback()

    db_session.add(
        Account(
            id=7302,
            user_id=USER_ID,
            provider="projectx",
            external_id="7302",
            trade_data_source="csv_import",
            account_state="ACTIVE",
        )
    )
    db_session.commit()
    db_session.add(
        ProjectXTradeEvent(
            user_id=USER_ID,
            account_id=7302,
            account_row_id=7302,
            account_external_id="7302",
            contract_id="MNQU6",
            symbol="MNQ",
            side="BUY",
            size=1,
            price=1,
            trade_timestamp=datetime.now(timezone.utc),
            fees=0,
            order_id="cross-owner",
            source_trade_id="cross-owner",
            import_batch_id=confirmed["import_id"],
        )
    )
    with pytest.raises(IntegrityError):
        db_session.commit()
    db_session.rollback()


def test_import_logs_are_structured_and_redacted(db_session, caplog):
    caplog.set_level(logging.INFO, logger="app.services.trade_imports")
    sensitive_id = "SENSITIVE-TRADE-ID"
    sensitive_filename = "private-account-history.csv"
    content = _csv_bytes([_trade_row(Id=sensitive_id, PnL="987.65")])

    preview, _ = _confirm(db_session, content, filename=sensitive_filename)

    output = "\n".join(record.getMessage() for record in caplog.records)
    assert "trade_import_preview" in output
    assert "trade_import_confirm" in output
    assert "parse_ms" in output
    assert "dedupe_ms" in output
    assert "commit_ms" in output
    assert sensitive_id not in output
    assert sensitive_filename not in output
    assert preview["preview_token"] not in output
    assert preview["file_sha256"] not in output
    assert "987.65" not in output


def test_import_failure_logs_category_and_timing_without_sensitive_values(db_session, caplog):
    caplog.set_level(logging.INFO, logger="app.services.trade_imports")
    sensitive_id = "FAILURE-SENSITIVE-ID"
    sensitive_filename = "failure-private-history.csv"

    with pytest.raises(TradeImportValidationError):
        _preview(
            db_session,
            _csv_bytes([_trade_row(Id=sensitive_id, Size="1.5", PnL="876.54")]),
            filename=sensitive_filename,
        )

    output = "\n".join(record.getMessage() for record in caplog.records)
    assert '"event":"trade_import_preview"' in output
    assert '"outcome":"validation:invalid_rows"' in output
    assert '"failure_phase":"parse"' in output
    assert '"error_rows":1' in output
    assert '"parse_ms":' in output
    assert '"total_ms":' in output
    assert sensitive_id not in output
    assert sensitive_filename not in output
    assert USER_ID not in output
    assert str(ACCOUNT_ID) not in output
    assert "876.54" not in output


def test_confirm_failure_logs_counts_phase_and_timing_without_sensitive_values(
    db_session,
    caplog,
):
    caplog.set_level(logging.INFO, logger="app.services.trade_imports")
    sensitive_id = "CONFIRM-FAILURE-SENSITIVE-ID"
    first = _trade_row(Id=sensitive_id)
    conflicting = _trade_row(Id=sensitive_id, PnL="765.43")
    preview = _preview(
        db_session,
        _csv_bytes([first, conflicting]),
        filename="confirm-failure-private.csv",
    )
    caplog.clear()

    with pytest.raises(TradeImportValidationError):
        confirm_trade_import(
            db_session,
            user_id=USER_ID,
            account_id=ACCOUNT_ID,
            preview_token=preview["preview_token"],
        )

    output = "\n".join(record.getMessage() for record in caplog.records)
    assert '"event":"trade_import_confirm"' in output
    assert '"outcome":"validation:import_conflicts_unresolved"' in output
    assert '"failure_phase":"preflight"' in output
    assert '"total_rows":2' in output
    assert '"new_rows":1' in output
    assert '"conflict_rows":1' in output
    assert '"total_ms":' in output
    assert sensitive_id not in output
    assert preview["preview_token"] not in output
    assert USER_ID not in output
    assert str(ACCOUNT_ID) not in output
    assert "765.43" not in output


@pytest.mark.parametrize("file_type", ["csv", "xlsx"])
def test_5000_row_import_path_preserves_counts(file_type, db_session):
    rows = [_trade_row(Id=str(3_000_000_000 + index)) for index in range(5_000)]
    content = _csv_bytes(rows) if file_type == "csv" else _xlsx_bytes(rows)
    started = perf_counter()
    preview, confirmed = _confirm(
        db_session,
        content,
        filename=f"five-thousand.{file_type}",
    )
    elapsed_seconds = perf_counter() - started
    assert preview["total_rows"] == 5_000
    assert preview["new_rows"] == 5_000
    assert confirmed["inserted_rows"] == 5_000
    assert db_session.query(ProjectXTradeEvent).count() == 5_000
    assert elapsed_seconds < MAX_5000_ROW_IMPORT_SECONDS, (
        f"{file_type} 5,000-row preview+confirm took {elapsed_seconds:.2f}s; "
        f"budget is {MAX_5000_ROW_IMPORT_SECONDS:.0f}s"
    )
