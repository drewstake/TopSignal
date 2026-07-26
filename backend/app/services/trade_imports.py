from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import math
import re
import secrets
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from time import perf_counter
from typing import Any, Callable, Iterable, Sequence
from zipfile import BadZipFile, ZipFile

from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..models import Account, ProjectXTradeEvent, TradeImportBatch, TradeImportPreview
from .instruments import normalize_symbol_key
from .projectx_accounts import ACCOUNT_PROVIDER, TRADE_DATA_SOURCE_CSV_IMPORT
from .trading_day import TRADING_TZ, trading_day_date


MAX_TRADE_IMPORT_BYTES = 10 * 1024 * 1024
MAX_TRADE_IMPORT_ROWS = 5_000
MAX_REPORTED_ROW_ERRORS = 100
MAX_SOURCE_TRADE_ID_CHARS = 255
MAX_EXCEL_ARCHIVE_ENTRIES = 10_000
MAX_EXCEL_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_FUTURES_QUANTITY = 10_000
TRADE_IMPORT_PREVIEW_TTL = timedelta(minutes=30)
TRADE_IMPORT_PREVIEW_RETENTION = timedelta(days=7)
TRADE_IMPORT_MANIFEST_VERSION = 1

_TRADING_TZ = TRADING_TZ
_MONEY_QUANT = Decimal("0.01")
_STORAGE_QUANT = Decimal("0.000001")
_STORAGE_MAX = Decimal("999999999999.999999")
_CSV_EXTENSIONS = {".csv"}
_EXCEL_EXTENSIONS = {".xlsx"}

logger = logging.getLogger(__name__)

_COLUMN_LABELS = {
    "source_trade_id": "Id",
    "contract_name": "ContractName",
    "entered_at": "EnteredAt",
    "exited_at": "ExitedAt",
    "entry_price": "EntryPrice",
    "exit_price": "ExitPrice",
    "fees": "Fees",
    "pnl": "PnL",
    "size": "Size",
    "direction": "Type",
    "trade_day": "TradeDay",
    "duration": "TradeDuration",
    "commissions": "Commissions",
}

_COLUMN_ALIASES = {
    "source_trade_id": {
        "id",
        "tradeid",
        "tradeidentifier",
        "executionid",
        "executionidentifier",
    },
    "contract_name": {
        "contract",
        "contractname",
        "contractsymbol",
        "instrument",
        "instrumentname",
        "symbol",
    },
    "entered_at": {
        "enteredat",
        "entryat",
        "entrytime",
        "entrytimestamp",
        "openedat",
        "opentime",
        "opentimestamp",
    },
    "exited_at": {
        "exitedat",
        "exitat",
        "exittime",
        "exittimestamp",
        "closedat",
        "closetime",
        "closetimestamp",
    },
    "entry_price": {
        "entryprice",
        "averageentryprice",
        "avgentryprice",
        "openprice",
    },
    "exit_price": {
        "exitprice",
        "averageexitprice",
        "avgexitprice",
        "closeprice",
    },
    "fees": {
        "fee",
        "fees",
        "brokerfee",
        "brokerfees",
        "exchangefee",
        "exchangefees",
        "noncommissionfees",
    },
    "pnl": {
        "pnl",
        "pandl",
        "profitandloss",
        "profitloss",
        "realizedpnl",
        "grosspnl",
    },
    "size": {
        "size",
        "quantity",
        "qty",
        "contracts",
        "contractquantity",
    },
    "direction": {
        "type",
        "tradetype",
        "direction",
        "tradedirection",
        "side",
        "positionside",
    },
    "trade_day": {
        "tradeday",
        "tradedate",
        "sessiondate",
        "tradingday",
        "tradingdate",
    },
    "duration": {
        "tradeduration",
        "duration",
        "holdingtime",
        "holdduration",
    },
    "commissions": {
        "commission",
        "commissions",
        "brokercommission",
        "brokercommissions",
        "topstepcommission",
        "topstepcommissions",
    },
}

_REQUIRED_COLUMNS = (
    "source_trade_id",
    "contract_name",
    "entered_at",
    "exited_at",
    "entry_price",
    "exit_price",
    "fees",
    "pnl",
    "size",
    "direction",
    "trade_day",
    "commissions",
)

_DATETIME_FORMATS = (
    "%m/%d/%Y %H:%M:%S.%f %z",
    "%m/%d/%Y %H:%M:%S %z",
    "%m/%d/%Y %I:%M:%S.%f %p %z",
    "%m/%d/%Y %I:%M:%S %p %z",
    "%Y-%m-%d %H:%M:%S.%f %z",
    "%Y-%m-%d %H:%M:%S %z",
    "%m/%d/%Y %H:%M:%S.%f",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %I:%M:%S.%f %p",
    "%m/%d/%Y %I:%M:%S %p",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
)

_DATE_FORMATS = (
    "%m/%d/%Y",
    "%Y-%m-%d",
    "%m-%d-%Y",
)


class TradeImportValidationError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        missing_columns: Sequence[str] | None = None,
        row_errors: Sequence[dict[str, Any]] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = str(code)
        self.message = str(message)
        self.missing_columns = list(missing_columns or [])
        self.row_errors = [dict(error) for error in (row_errors or [])]

    def as_detail(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "message": self.message,
            "missing_columns": list(self.missing_columns),
            "row_errors": [dict(error) for error in self.row_errors],
        }

    def as_dict(self) -> dict[str, Any]:
        return self.as_detail()


@dataclass(frozen=True)
class _SourceRow:
    row_number: int
    values: tuple[Any, ...]
    date_only_indexes: frozenset[int] = frozenset()


@dataclass(frozen=True)
class _ExcelCellValue:
    value: Any
    date_only: bool = False


@dataclass(frozen=True)
class _ParsedTrade:
    row_number: int
    source_trade_id: str
    contract_name: str
    symbol: str
    entered_at: datetime
    exited_at: datetime
    entry_price: Decimal
    exit_price: Decimal
    fees: Decimal
    commissions: Decimal
    gross_pnl: Decimal
    size: Decimal
    direction: str
    closing_side: str
    trade_day: date
    duration: str | None
    raw_columns: dict[str, Any]

    @property
    def net_pnl(self) -> Decimal:
        return self.gross_pnl - self.fees - self.commissions

    def preview_payload(
        self,
        *,
        status: str,
        conflict: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        payload = {
            "row_number": self.row_number,
            "source_trade_id": self.source_trade_id,
            "contract_name": self.contract_name,
            "symbol": self.symbol,
            "entered_at": self.entered_at,
            "exited_at": self.exited_at,
            "entry_price": _decimal_float(self.entry_price),
            "exit_price": _decimal_float(self.exit_price),
            "fees": _money_float(self.fees),
            "commissions": _money_float(self.commissions),
            "gross_pnl": _money_float(self.gross_pnl),
            "net_pnl": _money_float(self.net_pnl),
            "size": _decimal_float(self.size),
            "direction": self.direction,
            "trade_day": self.trade_day,
            "duration": self.duration,
            "status": status,
        }
        if conflict is not None:
            payload["conflict"] = conflict
        return payload


@dataclass(frozen=True)
class _ParsedFile:
    source_file_name: str
    file_sha256: str
    trades: tuple[_ParsedTrade, ...]


@dataclass(frozen=True)
class _PreparedImport:
    parsed: _ParsedFile
    decisions: tuple["_TradeDecision", ...]
    existing_batch: TradeImportBatch | None
    dedupe_snapshot: str


@dataclass(frozen=True)
class _TradeDecision:
    status: str
    identity_kind: str
    identity_value: str
    existing_event_ids: tuple[int, ...] = ()
    existing_fingerprints: tuple[str, ...] = ()
    conflict: dict[str, Any] | None = None


def preview_trade_import(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    filename: str,
    content: bytes,
    account_row_id: int | None = None,
) -> dict[str, Any]:
    overall_started = perf_counter()
    phase_started = overall_started
    failure_phase = "account_validation"
    parse_ms = None
    dedupe_ms = None
    persist_ms = None
    total_rows = None
    new_rows = None
    duplicate_rows = None
    conflict_rows = None
    try:
        account = _resolve_import_account(
            db,
            user_id=str(user_id),
            account_id=int(account_id),
            account_row_id=account_row_id,
        )
        failure_phase = "parse"
        phase_started = perf_counter()
        parsed = _parse_file(filename=filename, content=content)
        parse_ms = _elapsed_ms(phase_started)
        total_rows = len(parsed.trades)

        failure_phase = "dedupe"
        phase_started = perf_counter()
        prepared = _prepare_parsed_import(
            db,
            user_id=str(user_id),
            account_id=int(account_id),
            parsed=parsed,
        )
        dedupe_ms = _elapsed_ms(phase_started)

        decisions = prepared.decisions
        duplicate_rows = sum(1 for decision in decisions if decision.status == "duplicate")
        conflict_rows = sum(1 for decision in decisions if decision.status == "conflict")
        new_trades = [
            trade
            for trade, decision in zip(parsed.trades, decisions)
            if decision.status == "new"
        ]
        new_rows = len(new_trades)
        preview_rows = [
            trade.preview_payload(status=decision.status, conflict=decision.conflict)
            for trade, decision in zip(parsed.trades, decisions)
        ]

        now = datetime.now(timezone.utc)
        preview_token = secrets.token_urlsafe(32)
        staged = TradeImportPreview(
            token_hash=_token_hash(preview_token),
            user_id=str(user_id),
            account_id=int(account_id),
            account_row_id=int(account.id),
            account_external_id=str(account.external_id),
            source_file_name=parsed.source_file_name,
            file_sha256=parsed.file_sha256,
            manifest_version=TRADE_IMPORT_MANIFEST_VERSION,
            normalized_manifest=[_serialize_manifest_trade(trade) for trade in parsed.trades],
            preview_rows=_json_compatible(preview_rows),
            dedupe_snapshot=prepared.dedupe_snapshot,
            total_rows=len(parsed.trades),
            new_rows=len(new_trades),
            duplicate_rows=duplicate_rows,
            conflict_rows=conflict_rows,
            status="conflict" if conflict_rows else "pending",
            outcome_code="identity_conflict" if conflict_rows else None,
            created_at=now,
            updated_at=now,
            expires_at=now + TRADE_IMPORT_PREVIEW_TTL,
            retention_until=now + TRADE_IMPORT_PREVIEW_RETENTION,
        )
        failure_phase = "persist"
        phase_started = perf_counter()
        _cleanup_expired_previews(db, now=now)
        db.add(staged)
        db.commit()
        persist_ms = _elapsed_ms(phase_started)

        payload = {
            "preview_token": preview_token,
            "expires_at": staged.expires_at,
            "source_file_name": parsed.source_file_name,
            "file_sha256": parsed.file_sha256,
            "total_rows": len(parsed.trades),
            "new_rows": len(new_trades),
            "duplicate_rows": duplicate_rows,
            "conflict_rows": conflict_rows,
            "summary": _summarize_trades(new_trades),
            "trades": preview_rows,
        }
        _log_import_outcome(
            "preview",
            outcome="conflict" if conflict_rows else "ready",
            total_rows=len(parsed.trades),
            new_rows=len(new_trades),
            duplicate_rows=duplicate_rows,
            conflict_rows=conflict_rows,
            parse_ms=parse_ms,
            dedupe_ms=dedupe_ms,
            persist_ms=persist_ms,
            total_ms=_elapsed_ms(overall_started),
        )
        return payload
    except Exception as exc:
        db.rollback()
        if failure_phase == "parse" and parse_ms is None:
            parse_ms = _elapsed_ms(phase_started)
        elif failure_phase == "dedupe" and dedupe_ms is None:
            dedupe_ms = _elapsed_ms(phase_started)
        elif failure_phase == "persist" and persist_ms is None:
            persist_ms = _elapsed_ms(phase_started)
        _log_import_outcome(
            "preview",
            outcome=_safe_failure_category(exc),
            failure_phase=failure_phase,
            error_rows=(
                len(exc.row_errors)
                if isinstance(exc, TradeImportValidationError)
                else None
            ),
            total_rows=total_rows,
            new_rows=new_rows,
            duplicate_rows=duplicate_rows,
            conflict_rows=conflict_rows,
            parse_ms=parse_ms,
            dedupe_ms=dedupe_ms,
            persist_ms=persist_ms,
            total_ms=_elapsed_ms(overall_started),
        )
        raise


def confirm_trade_import(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    preview_token: str,
    account_row_id: int | None = None,
) -> dict[str, Any]:
    overall_started = perf_counter()
    resolved_user_id = str(user_id)
    resolved_account_id = int(account_id)
    phase_started = overall_started
    failure_phase = "account_validation"
    dedupe_ms = None
    commit_ms = None
    total_rows = None
    new_rows = None
    duplicate_rows = None
    conflict_rows = None
    try:
        account = _resolve_import_account(
            db,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            account_row_id=account_row_id,
        )
        now = datetime.now(timezone.utc)
        token_digest = _token_hash(_validated_preview_token(preview_token))
        staged = _get_staged_preview(
            db,
            token_hash=token_digest,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            account_row_id=int(account.id),
        )
        if staged is None:
            raise TradeImportValidationError(
                "preview_not_found",
                "The import preview was not found for this account. Preview the file again.",
            )
        total_rows = int(staged.total_rows)
        new_rows = int(staged.new_rows)
        duplicate_rows = int(staged.duplicate_rows)
        conflict_rows = int(staged.conflict_rows)
        failure_phase = "preflight"
        if staged.status == "committed" and staged.import_batch_id is not None:
            batch = _owned_batch_for_preview(db, staged)
            if batch is None:
                raise TradeImportValidationError(
                    "import_status_inconsistent",
                    "The saved import outcome could not be verified.",
                )
            return _serialize_existing_batch(batch)
        if _as_utc(staged.expires_at) <= now or staged.status == "expired":
            expired = _expire_preview_conditionally(
                db,
                preview_id=int(staged.id),
                now=now,
            )
            db.commit()
            db.expire_all()
            staged = _get_staged_preview(
                db,
                token_hash=token_digest,
                user_id=resolved_user_id,
                account_id=resolved_account_id,
                account_row_id=int(account.id),
            )
            if staged is not None and staged.status == "committed" and staged.import_batch_id is not None:
                batch = _owned_batch_for_preview(db, staged)
                if batch is not None:
                    return _serialize_existing_batch(batch)
            if expired or (staged is not None and staged.status == "expired"):
                raise TradeImportValidationError(
                    "preview_expired",
                    "The import preview expired. Preview the file again before importing.",
                )
            if staged is not None and staged.status == "confirming":
                raise TradeImportValidationError(
                    "confirmation_in_progress",
                    "This import confirmation is already in progress. Check its status before retrying.",
                )
            if staged is None:
                raise TradeImportValidationError(
                    "preview_not_found",
                    "The import preview was not found for this account. Preview the file again.",
                )
        if staged.status == "conflict" or int(staged.conflict_rows) > 0:
            raise TradeImportValidationError(
                "import_conflicts_unresolved",
                "The preview contains conflicting trade identities and cannot be imported.",
            )
        if staged.status == "stale":
            raise TradeImportValidationError(
                "preview_stale",
                "Account data changed after this preview. Preview the file again.",
            )
        if staged.status == "failed":
            raise TradeImportValidationError(
                "preview_failed",
                "The previous confirmation failed. Preview the file again.",
            )

        # A conditional write is both the state transition and the concurrency
        # gate. PostgreSQL locks the row until this transaction completes;
        # SQLite serializes the write. A second confirmation observes either a
        # committed outcome or a non-pending preview and cannot insert twice.
        failure_phase = "claim"
        phase_started = perf_counter()
        claimed = (
            db.query(TradeImportPreview)
            .filter(TradeImportPreview.id == staged.id)
            .filter(TradeImportPreview.status == "pending")
            .update(
                {
                    TradeImportPreview.status: "confirming",
                    TradeImportPreview.updated_at: now,
                },
                synchronize_session=False,
            )
        )
        if claimed != 1:
            db.expire_all()
            latest = _get_staged_preview(
                db,
                token_hash=token_digest,
                user_id=resolved_user_id,
                account_id=resolved_account_id,
                account_row_id=int(account.id),
            )
            if latest is not None and latest.status == "committed" and latest.import_batch_id is not None:
                batch = _owned_batch_for_preview(db, latest)
                if batch is not None:
                    return _serialize_existing_batch(batch)
            raise TradeImportValidationError(
                "confirmation_in_progress",
                "This import confirmation is already in progress. Check its status before retrying.",
            )

        db.expire(staged)
        staged = db.query(TradeImportPreview).filter(TradeImportPreview.id == staged.id).one()
        parsed = _parsed_file_from_stage(staged)
        failure_phase = "dedupe"
        phase_started = perf_counter()
        prepared = _prepare_parsed_import(
            db,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            parsed=parsed,
        )
        dedupe_ms = _elapsed_ms(phase_started)
        new_rows = sum(1 for decision in prepared.decisions if decision.status == "new")
        duplicate_rows = sum(
            1 for decision in prepared.decisions if decision.status == "duplicate"
        )
        conflict_rows = sum(
            1 for decision in prepared.decisions if decision.status == "conflict"
        )
        if prepared.dedupe_snapshot != staged.dedupe_snapshot:
            staged.status = "stale"
            staged.outcome_code = "preview_stale"
            staged.updated_at = now
            _scrub_staged_manifest(staged)
            db.commit()
            raise TradeImportValidationError(
                "preview_stale",
                "Account data changed after this preview. Preview the file again.",
            )
        if any(decision.status == "conflict" for decision in prepared.decisions):
            staged.status = "conflict"
            staged.outcome_code = "identity_conflict"
            staged.updated_at = now
            db.commit()
            raise TradeImportValidationError(
                "import_conflicts_unresolved",
                "The preview contains conflicting trade identities and cannot be imported.",
            )

        if prepared.existing_batch is not None:
            _mark_preview_committed(staged, prepared.existing_batch, now=now)
            db.commit()
            result = _serialize_existing_batch(prepared.existing_batch)
            _log_import_outcome(
                "confirm",
                outcome="idempotent",
                total_rows=result["total_rows"],
                new_rows=0,
                duplicate_rows=result["duplicate_rows"],
                conflict_rows=0,
                dedupe_ms=dedupe_ms,
                commit_ms=0.0,
                total_ms=_elapsed_ms(overall_started),
            )
            return result

        failure_phase = "commit"
        phase_started = perf_counter()
        result = _persist_staged_import(
            db,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            account=account,
            prepared=prepared,
            staged=staged,
            now=now,
        )
        commit_ms = _elapsed_ms(phase_started)
        _log_import_outcome(
            "confirm",
            outcome="committed",
            total_rows=result["total_rows"],
            new_rows=result["inserted_rows"],
            duplicate_rows=result["duplicate_rows"],
            conflict_rows=0,
            dedupe_ms=dedupe_ms,
            commit_ms=commit_ms,
            total_ms=_elapsed_ms(overall_started),
        )
        return result
    except IntegrityError as exc:
        if failure_phase == "dedupe" and dedupe_ms is None:
            dedupe_ms = _elapsed_ms(phase_started)
        elif failure_phase == "commit" and commit_ms is None:
            commit_ms = _elapsed_ms(phase_started)
        db.rollback()
        # A same-token or same-file concurrent request may have committed while
        # this transaction was waiting. Recover that durable outcome rather
        # than reporting a false failure.
        recovered = _recover_concurrent_commit(
            db,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            preview_token=preview_token,
            account_row_id=account_row_id,
        )
        if recovered is not None:
            _log_import_outcome(
                "confirm",
                outcome="concurrent_idempotent",
                total_rows=recovered["total_rows"],
                new_rows=0,
                duplicate_rows=recovered["duplicate_rows"],
                conflict_rows=0,
                failure_phase=failure_phase,
                dedupe_ms=dedupe_ms,
                commit_ms=commit_ms,
                total_ms=_elapsed_ms(overall_started),
            )
            return recovered
        _mark_preview_stale_after_integrity_conflict(
            db,
            user_id=resolved_user_id,
            account_id=resolved_account_id,
            preview_token=preview_token,
            account_row_id=account_row_id,
        )
        _log_import_outcome(
            "confirm",
            outcome="preview_stale",
            failure_phase=failure_phase,
            total_rows=total_rows,
            new_rows=new_rows,
            duplicate_rows=duplicate_rows,
            conflict_rows=conflict_rows,
            dedupe_ms=dedupe_ms,
            commit_ms=commit_ms,
            total_ms=_elapsed_ms(overall_started),
        )
        raise TradeImportValidationError(
            "preview_stale",
            "Account data changed while this import was being confirmed. Preview it again.",
        ) from exc
    except Exception as exc:
        if failure_phase == "dedupe" and dedupe_ms is None:
            dedupe_ms = _elapsed_ms(phase_started)
        elif failure_phase == "commit" and commit_ms is None:
            commit_ms = _elapsed_ms(phase_started)
        if db.in_transaction():
            db.rollback()
        _log_import_outcome(
            "confirm",
            outcome=_safe_failure_category(exc),
            failure_phase=failure_phase,
            error_rows=(
                len(exc.row_errors)
                if isinstance(exc, TradeImportValidationError)
                else None
            ),
            total_rows=total_rows,
            new_rows=new_rows,
            duplicate_rows=duplicate_rows,
            conflict_rows=conflict_rows,
            dedupe_ms=dedupe_ms,
            commit_ms=commit_ms,
            total_ms=_elapsed_ms(overall_started),
        )
        raise


def get_trade_import_status(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    preview_token: str,
    account_row_id: int | None = None,
) -> dict[str, Any]:
    account = _resolve_import_account(
        db,
        user_id=str(user_id),
        account_id=int(account_id),
        account_row_id=account_row_id,
        allow_archived=True,
    )
    token_digest = _token_hash(_validated_preview_token(preview_token))
    staged = _get_staged_preview(
        db,
        token_hash=token_digest,
        user_id=str(user_id),
        account_id=int(account_id),
        account_row_id=int(account.id),
    )
    if staged is None:
        raise TradeImportValidationError(
            "preview_not_found",
            "The import preview was not found for this account.",
        )
    now = datetime.now(timezone.utc)
    # ``confirming`` is only an in-transaction claim and is never committed on
    # the successful path. If a legacy or interrupted deployment left that
    # state durable, no live confirmation can still own it once this query can
    # read the row, so returning it to pending is safe and makes recovery
    # deterministic after a restart.
    if staged.status == "confirming":
        (
            db.query(TradeImportPreview)
            .filter(TradeImportPreview.id == staged.id)
            .filter(TradeImportPreview.status == "confirming")
            .update(
                {
                    TradeImportPreview.status: "pending",
                    TradeImportPreview.outcome_code: "confirmation_retryable",
                    TradeImportPreview.updated_at: now,
                },
                synchronize_session=False,
            )
        )
        db.commit()
        db.expire_all()
        staged = _get_staged_preview(
            db,
            token_hash=token_digest,
            user_id=str(user_id),
            account_id=int(account_id),
            account_row_id=int(account.id),
        )
        if staged is None:
            raise TradeImportValidationError(
                "preview_not_found",
                "The import preview was not found for this account.",
            )
    if staged.status in {"pending", "conflict"} and _as_utc(staged.expires_at) <= now:
        _expire_preview_conditionally(
            db,
            preview_id=int(staged.id),
            now=now,
        )
        db.commit()
        db.expire_all()
        staged = _get_staged_preview(
            db,
            token_hash=token_digest,
            user_id=str(user_id),
            account_id=int(account_id),
            account_row_id=int(account.id),
        )
        if staged is None:
            raise TradeImportValidationError(
                "preview_not_found",
                "The import preview was not found for this account.",
            )
    result = None
    if staged.status == "committed" and staged.import_batch_id is not None:
        batch = _owned_batch_for_preview(db, staged)
        result = _serialize_existing_batch(batch) if batch is not None else None
    return {
        "status": staged.status,
        "confirmation_retryable": staged.status == "pending" and account.archived_at is None,
        "outcome_code": staged.outcome_code,
        "source_file_name": staged.source_file_name,
        "created_at": staged.created_at,
        "expires_at": staged.expires_at,
        "confirmed_at": staged.confirmed_at,
        "total_rows": int(staged.total_rows),
        "new_rows": int(staged.new_rows),
        "duplicate_rows": int(staged.duplicate_rows),
        "conflict_rows": int(staged.conflict_rows),
        "result": result,
    }


def cleanup_trade_import_previews(
    db: Session,
    *,
    now: datetime | None = None,
) -> dict[str, int]:
    """Scrub expired manifests and delete preview metadata past retention."""
    resolved_now = _as_utc(now or datetime.now(timezone.utc))
    try:
        expired_previews, deleted_previews = _cleanup_expired_previews(
            db,
            now=resolved_now,
        )
        db.commit()
    except Exception:
        db.rollback()
        raise
    _log_import_outcome(
        "cleanup",
        outcome="complete",
        expired_previews=expired_previews,
        deleted_previews=deleted_previews,
    )
    return {
        "expired_previews": expired_previews,
        "deleted_previews": deleted_previews,
    }


def _prepare_parsed_import(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    parsed: _ParsedFile,
) -> _PreparedImport:
    source_map, fallback_map = _load_existing_identity_candidates(
        db,
        user_id=user_id,
        account_id=account_id,
        trades=parsed.trades,
    )

    decisions: list[_TradeDecision] = []
    seen_in_file: dict[str, _ParsedTrade] = {}
    for trade in parsed.trades:
        decision = _classify_trade_identity(
            trade,
            prior_in_file=seen_in_file.get(trade.source_trade_id),
            source_candidates=source_map.get(trade.source_trade_id, ()),
            fallback_candidates=fallback_map.get(
                (trade.source_trade_id, _timestamp_key(trade.exited_at)),
                (),
            ),
        )
        decisions.append(decision)
        seen_in_file.setdefault(trade.source_trade_id, trade)

    existing_batch = (
        db.query(TradeImportBatch)
        .filter(TradeImportBatch.user_id == user_id)
        .filter(TradeImportBatch.account_id == account_id)
        .filter(TradeImportBatch.file_sha256 == parsed.file_sha256)
        .first()
    )
    return _PreparedImport(
        parsed=parsed,
        decisions=tuple(decisions),
        existing_batch=existing_batch,
        dedupe_snapshot=_dedupe_snapshot(decisions),
    )


def _persist_staged_import(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    account: Account,
    prepared: _PreparedImport,
    staged: TradeImportPreview,
    now: datetime,
) -> dict[str, Any]:
    parsed = prepared.parsed
    imported_at = now
    duplicate_rows = sum(1 for decision in prepared.decisions if decision.status == "duplicate")
    inserted_rows = sum(1 for decision in prepared.decisions if decision.status == "new")
    if any(decision.status == "conflict" for decision in prepared.decisions):
        raise TradeImportValidationError(
            "import_conflicts_unresolved",
            "The preview contains conflicting trade identities and cannot be imported.",
        )
    batch = TradeImportBatch(
        user_id=user_id,
        account_id=account_id,
        account_row_id=int(account.id),
        account_external_id=str(account.external_id),
        source_file_name=parsed.source_file_name,
        file_sha256=parsed.file_sha256,
        imported_at=imported_at,
        total_rows=len(parsed.trades),
        inserted_rows=inserted_rows,
        duplicate_rows=duplicate_rows,
    )

    db.add(batch)
    db.flush()

    events = [
        ProjectXTradeEvent(
            user_id=user_id,
            account_id=account_id,
            account_row_id=int(account.id),
            account_external_id=str(account.external_id),
            contract_id=trade.contract_name,
            symbol=trade.symbol,
            side=trade.closing_side,
            size=trade.size,
            price=trade.exit_price,
            trade_timestamp=trade.exited_at,
            fees=trade.fees,
            commissions=trade.commissions,
            fee_scope="round_turn",
            pnl=trade.gross_pnl,
            trade_date=trade.trade_day,
            entry_timestamp=trade.entered_at,
            entry_price=trade.entry_price,
            order_id=trade.source_trade_id,
            source_trade_id=trade.source_trade_id,
            status="IMPORTED",
            raw_payload={
                "source": "topstep_trade_export",
                "manifest_version": TRADE_IMPORT_MANIFEST_VERSION,
                "row_number": trade.row_number,
            },
            import_batch_id=int(batch.id),
        )
        for trade, decision in zip(parsed.trades, prepared.decisions)
        if decision.status == "new"
    ]
    if events:
        db.bulk_save_objects(events)

    _mark_preview_committed(staged, batch, now=now)
    db.commit()

    return {
        "import_id": int(batch.id),
        "source_file_name": parsed.source_file_name,
        "imported_at": imported_at,
        "total_rows": len(parsed.trades),
        "inserted_rows": inserted_rows,
        "duplicate_rows": duplicate_rows,
    }


def _serialize_existing_batch(batch: TradeImportBatch) -> dict[str, Any]:
    imported_at = batch.imported_at
    if imported_at.tzinfo is None:
        imported_at = imported_at.replace(tzinfo=timezone.utc)
    else:
        imported_at = imported_at.astimezone(timezone.utc)
    return {
        "import_id": int(batch.id),
        "source_file_name": str(batch.source_file_name),
        "imported_at": imported_at,
        "total_rows": int(batch.total_rows),
        "inserted_rows": int(batch.inserted_rows),
        "duplicate_rows": int(batch.duplicate_rows),
    }


def _resolve_import_account(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    account_row_id: int | None,
    allow_archived: bool = False,
) -> Account:
    query = (
        db.query(Account)
        .filter(Account.user_id == str(user_id))
        .filter(Account.provider == ACCOUNT_PROVIDER)
        .filter(Account.external_id == str(account_id))
    )
    if account_row_id is not None:
        query = query.filter(Account.id == int(account_row_id))
    account = query.one_or_none()
    if account is None:
        raise TradeImportValidationError(
            "import_account_not_found",
            "The Live account was not found.",
        )
    if account.trade_data_source != TRADE_DATA_SOURCE_CSV_IMPORT:
        raise TradeImportValidationError(
            "trade_import_requires_csv_import_account",
            "Trade imports require a Topstep Live CSV account.",
        )
    if not allow_archived and account.archived_at is not None:
        raise TradeImportValidationError(
            "trade_import_account_archived",
            "Restore this Live account before importing trades.",
        )
    return account


def _validated_preview_token(value: str) -> str:
    token = str(value or "").strip()
    if not 32 <= len(token) <= 200 or not re.fullmatch(r"[A-Za-z0-9_-]+", token):
        raise TradeImportValidationError(
            "invalid_preview_token",
            "The import preview token is invalid. Preview the file again.",
        )
    return token


def _token_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _get_staged_preview(
    db: Session,
    *,
    token_hash: str,
    user_id: str,
    account_id: int,
    account_row_id: int,
) -> TradeImportPreview | None:
    return (
        db.query(TradeImportPreview)
        .filter(TradeImportPreview.token_hash == token_hash)
        .filter(TradeImportPreview.user_id == user_id)
        .filter(TradeImportPreview.account_id == account_id)
        .filter(TradeImportPreview.account_row_id == account_row_id)
        .one_or_none()
    )


def _owned_batch_for_preview(
    db: Session,
    preview: TradeImportPreview,
) -> TradeImportBatch | None:
    if preview.import_batch_id is None:
        return None
    return (
        db.query(TradeImportBatch)
        .filter(TradeImportBatch.id == preview.import_batch_id)
        .filter(TradeImportBatch.user_id == preview.user_id)
        .filter(TradeImportBatch.account_id == preview.account_id)
        .filter(TradeImportBatch.account_row_id == preview.account_row_id)
        .filter(TradeImportBatch.account_external_id == preview.account_external_id)
        .one_or_none()
    )


def _mark_preview_committed(
    preview: TradeImportPreview,
    batch: TradeImportBatch,
    *,
    now: datetime,
) -> None:
    preview.status = "committed"
    preview.outcome_code = "committed"
    preview.import_batch_id = int(batch.id)
    preview.confirmed_at = now
    preview.updated_at = now
    _scrub_staged_manifest(preview)


def _expire_preview_conditionally(
    db: Session,
    *,
    preview_id: int,
    now: datetime,
) -> bool:
    updated = (
        db.query(TradeImportPreview)
        .filter(TradeImportPreview.id == preview_id)
        .filter(
            TradeImportPreview.status.in_(
                ("pending", "confirming", "conflict", "stale", "failed")
            )
        )
        .filter(TradeImportPreview.expires_at <= now)
        .update(
            {
                TradeImportPreview.status: "expired",
                TradeImportPreview.outcome_code: "preview_expired",
                TradeImportPreview.updated_at: now,
                TradeImportPreview.normalized_manifest: None,
                TradeImportPreview.preview_rows: None,
                TradeImportPreview.dedupe_snapshot: None,
            },
            synchronize_session=False,
        )
    )
    return int(updated) == 1


def _scrub_staged_manifest(preview: TradeImportPreview) -> None:
    preview.normalized_manifest = None
    preview.preview_rows = None
    preview.dedupe_snapshot = None


def _cleanup_expired_previews(db: Session, *, now: datetime) -> tuple[int, int]:
    # Keep the status/expiry predicates in the UPDATE itself. PostgreSQL
    # rechecks them after waiting on a concurrent confirmation's row lock, so
    # cleanup cannot overwrite a preview that became committed while blocked.
    expired = (
        db.query(TradeImportPreview)
        .filter(
            TradeImportPreview.status.in_(
                ("pending", "confirming", "conflict", "stale", "failed")
            )
        )
        .filter(TradeImportPreview.expires_at <= now)
        .update(
            {
                TradeImportPreview.status: "expired",
                TradeImportPreview.outcome_code: "preview_expired",
                TradeImportPreview.updated_at: now,
                TradeImportPreview.normalized_manifest: None,
                TradeImportPreview.preview_rows: None,
                TradeImportPreview.dedupe_snapshot: None,
            },
            synchronize_session=False,
        )
    )
    deleted = (
        db.query(TradeImportPreview)
        .filter(TradeImportPreview.retention_until <= now)
        .delete(synchronize_session=False)
    )
    return int(expired), int(deleted)


def _serialize_manifest_trade(trade: _ParsedTrade) -> dict[str, Any]:
    return {
        "row_number": trade.row_number,
        "source_trade_id": trade.source_trade_id,
        "contract_name": trade.contract_name,
        "symbol": trade.symbol,
        "entered_at": _as_utc(trade.entered_at).isoformat(),
        "exited_at": _as_utc(trade.exited_at).isoformat(),
        "entry_price": str(trade.entry_price),
        "exit_price": str(trade.exit_price),
        "fees": str(trade.fees),
        "commissions": str(trade.commissions),
        "gross_pnl": str(trade.gross_pnl),
        "size": str(trade.size),
        "direction": trade.direction,
        "closing_side": trade.closing_side,
        "trade_day": trade.trade_day.isoformat(),
        "duration": trade.duration,
    }


def _parsed_file_from_stage(preview: TradeImportPreview) -> _ParsedFile:
    if preview.manifest_version != TRADE_IMPORT_MANIFEST_VERSION:
        raise TradeImportValidationError(
            "preview_version_unsupported",
            "The import preview format changed. Preview the file again.",
        )
    manifest = preview.normalized_manifest
    if not isinstance(manifest, list) or not manifest:
        raise TradeImportValidationError(
            "preview_manifest_unavailable",
            "The staged import data is unavailable. Preview the file again.",
        )
    trades: list[_ParsedTrade] = []
    try:
        for item in manifest:
            if not isinstance(item, dict):
                raise ValueError("invalid manifest row")
            trades.append(
                _ParsedTrade(
                    row_number=int(item["row_number"]),
                    source_trade_id=str(item["source_trade_id"]),
                    contract_name=str(item["contract_name"]),
                    symbol=str(item["symbol"]),
                    entered_at=datetime.fromisoformat(str(item["entered_at"])),
                    exited_at=datetime.fromisoformat(str(item["exited_at"])),
                    entry_price=Decimal(str(item["entry_price"])),
                    exit_price=Decimal(str(item["exit_price"])),
                    fees=Decimal(str(item["fees"])),
                    commissions=Decimal(str(item["commissions"])),
                    gross_pnl=Decimal(str(item["gross_pnl"])),
                    size=Decimal(str(item["size"])),
                    direction=str(item["direction"]),
                    closing_side=str(item["closing_side"]),
                    trade_day=date.fromisoformat(str(item["trade_day"])),
                    duration=(str(item["duration"]) if item.get("duration") is not None else None),
                    raw_columns={},
                )
            )
    except (KeyError, TypeError, ValueError, InvalidOperation) as exc:
        raise TradeImportValidationError(
            "preview_manifest_invalid",
            "The staged import data is invalid. Preview the file again.",
        ) from exc
    return _ParsedFile(
        source_file_name=str(preview.source_file_name),
        file_sha256=str(preview.file_sha256),
        trades=tuple(trades),
    )


def _json_compatible(value: Any) -> Any:
    if isinstance(value, datetime):
        return _as_utc(value).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _json_compatible(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_compatible(item) for item in value]
    return value


def _load_existing_identity_candidates(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    trades: Sequence[_ParsedTrade],
) -> tuple[dict[str, tuple[ProjectXTradeEvent, ...]], dict[tuple[str, str], tuple[ProjectXTradeEvent, ...]]]:
    source_ids = list(dict.fromkeys(trade.source_trade_id for trade in trades))
    source_map: defaultdict[str, list[ProjectXTradeEvent]] = defaultdict(list)
    fallback_map: defaultdict[tuple[str, str], list[ProjectXTradeEvent]] = defaultdict(list)
    chunk_size = 400
    seen_event_ids: set[int] = set()
    for start in range(0, len(source_ids), chunk_size):
        chunk = source_ids[start : start + chunk_size]
        rows = (
            db.query(ProjectXTradeEvent)
            .filter(ProjectXTradeEvent.user_id == user_id)
            .filter(ProjectXTradeEvent.account_id == account_id)
            .filter(
                or_(
                    ProjectXTradeEvent.source_trade_id.in_(chunk),
                    ProjectXTradeEvent.order_id.in_(chunk),
                )
            )
            .all()
        )
        for row in rows:
            row_id = int(row.id)
            if row_id in seen_event_ids:
                continue
            seen_event_ids.add(row_id)
            if row.source_trade_id is not None and str(row.source_trade_id) in source_ids:
                source_map[str(row.source_trade_id)].append(row)
            if row.order_id is not None and row.trade_timestamp is not None and str(row.order_id) in source_ids:
                fallback_map[(str(row.order_id), _timestamp_key(row.trade_timestamp))].append(row)
    return (
        {key: tuple(sorted(rows, key=lambda row: int(row.id))) for key, rows in source_map.items()},
        {key: tuple(sorted(rows, key=lambda row: int(row.id))) for key, rows in fallback_map.items()},
    )


def _classify_trade_identity(
    trade: _ParsedTrade,
    *,
    prior_in_file: _ParsedTrade | None,
    source_candidates: Sequence[ProjectXTradeEvent],
    fallback_candidates: Sequence[ProjectXTradeEvent],
) -> _TradeDecision:
    identity_kind = "source_trade_id" if source_candidates or prior_in_file is not None else "order_exit"
    identity_value = (
        trade.source_trade_id
        if identity_kind == "source_trade_id"
        else f"{trade.source_trade_id}|{_timestamp_key(trade.exited_at)}"
    )
    incoming = _canonical_incoming_trade(
        trade,
        identity_kind=identity_kind,
        identity_value=identity_value,
    )

    if prior_in_file is not None:
        prior = _canonical_incoming_trade(
            prior_in_file,
            identity_kind="source_trade_id",
            identity_value=trade.source_trade_id,
        )
        differences = _economic_differences(prior, incoming)
        if differences:
            conflict = {
                "identity_kind": "source_trade_id",
                "identity_value": trade.source_trade_id,
                "reason": "repeated_id_mismatch",
                "stored_row_number": prior_in_file.row_number,
                "differences": differences,
            }
            return _TradeDecision(
                status="conflict",
                identity_kind="source_trade_id",
                identity_value=trade.source_trade_id,
                existing_fingerprints=(_canonical_fingerprint(prior),),
                conflict=conflict,
            )

    candidates_by_id: dict[int, ProjectXTradeEvent] = {}
    for candidate in (*source_candidates, *fallback_candidates):
        candidates_by_id[int(candidate.id)] = candidate
    candidates = [candidates_by_id[key] for key in sorted(candidates_by_id)]
    if len(candidates) > 1:
        conflict = {
            "identity_kind": identity_kind,
            "identity_value": identity_value,
            "reason": "ambiguous_stored_identity",
            "stored_event_ids": [int(candidate.id) for candidate in candidates],
            "differences": [],
        }
        return _TradeDecision(
            status="conflict",
            identity_kind=identity_kind,
            identity_value=identity_value,
            existing_event_ids=tuple(int(candidate.id) for candidate in candidates),
            existing_fingerprints=tuple(
                _canonical_fingerprint(
                    _canonical_stored_trade(
                        candidate,
                        identity_kind=identity_kind,
                        identity_value=identity_value,
                    )
                )
                for candidate in candidates
            ),
            conflict=conflict,
        )
    if candidates:
        candidate = candidates[0]
        stored = _canonical_stored_trade(
            candidate,
            identity_kind=identity_kind,
            identity_value=identity_value,
        )
        differences = _economic_differences(stored, incoming)
        fingerprint = _canonical_fingerprint(stored)
        if differences:
            conflict = {
                "identity_kind": identity_kind,
                "identity_value": identity_value,
                "reason": "stored_trade_mismatch",
                "stored_event_id": int(candidate.id),
                "differences": differences,
            }
            return _TradeDecision(
                status="conflict",
                identity_kind=identity_kind,
                identity_value=identity_value,
                existing_event_ids=(int(candidate.id),),
                existing_fingerprints=(fingerprint,),
                conflict=conflict,
            )
        return _TradeDecision(
            status="duplicate",
            identity_kind=identity_kind,
            identity_value=identity_value,
            existing_event_ids=(int(candidate.id),),
            existing_fingerprints=(fingerprint,),
        )
    if prior_in_file is not None:
        prior = _canonical_incoming_trade(
            prior_in_file,
            identity_kind="source_trade_id",
            identity_value=trade.source_trade_id,
        )
        return _TradeDecision(
            status="duplicate",
            identity_kind="source_trade_id",
            identity_value=trade.source_trade_id,
            existing_fingerprints=(_canonical_fingerprint(prior),),
        )
    return _TradeDecision(
        status="new",
        identity_kind=identity_kind,
        identity_value=identity_value,
    )


_ECONOMIC_FIELDS = (
    "entered_at",
    "exited_at",
    "contract",
    "symbol",
    "direction",
    "quantity",
    "entry_price",
    "exit_price",
    "gross_pnl",
    "fees",
    "commissions",
    "net_pnl",
    "trade_day",
)


def _canonical_incoming_trade(
    trade: _ParsedTrade,
    *,
    identity_kind: str,
    identity_value: str,
) -> dict[str, Any]:
    return {
        "identity_kind": identity_kind,
        "identity_value": identity_value,
        "entered_at": _timestamp_key(trade.entered_at),
        "exited_at": _timestamp_key(trade.exited_at),
        "contract": trade.contract_name.upper(),
        "symbol": trade.symbol.upper(),
        "direction": trade.direction,
        "quantity": _canonical_decimal(trade.size),
        "entry_price": _canonical_decimal(trade.entry_price),
        "exit_price": _canonical_decimal(trade.exit_price),
        "gross_pnl": _canonical_decimal(trade.gross_pnl),
        "fees": _canonical_decimal(trade.fees),
        "commissions": _canonical_decimal(trade.commissions),
        "net_pnl": _canonical_decimal(trade.net_pnl),
        "trade_day": trade.trade_day.isoformat(),
    }


def _canonical_stored_trade(
    event: ProjectXTradeEvent,
    *,
    identity_kind: str,
    identity_value: str,
) -> dict[str, Any]:
    gross = _decimal_or_none(event.pnl)
    fees = _decimal_or_none(event.fees)
    commissions = _decimal_or_none(event.commissions)
    net = None
    if gross is not None and fees is not None and commissions is not None:
        net = gross - fees - commissions
    closing_side = str(event.side or "").upper()
    direction = "Long" if closing_side == "SELL" else ("Short" if closing_side == "BUY" else None)
    return {
        "identity_kind": identity_kind,
        "identity_value": identity_value,
        "entered_at": _timestamp_key(event.entry_timestamp) if event.entry_timestamp is not None else None,
        "exited_at": _timestamp_key(event.trade_timestamp),
        "contract": str(event.contract_id or "").upper(),
        "symbol": (normalize_symbol_key(str(event.symbol or event.contract_id or "")) or str(event.symbol or "").upper()),
        "direction": direction,
        "quantity": _canonical_decimal(event.size),
        "entry_price": _canonical_decimal(event.entry_price) if event.entry_price is not None else None,
        "exit_price": _canonical_decimal(event.price),
        "gross_pnl": _canonical_decimal(gross) if gross is not None else None,
        "fees": _canonical_decimal(fees) if fees is not None else None,
        "commissions": _canonical_decimal(commissions) if commissions is not None else None,
        "net_pnl": _canonical_decimal(net) if net is not None else None,
        "trade_day": event.trade_date.isoformat() if event.trade_date is not None else None,
    }


def _economic_differences(stored: dict[str, Any], incoming: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "field": field,
            "stored": stored.get(field),
            "incoming": incoming.get(field),
        }
        for field in _ECONOMIC_FIELDS
        if stored.get(field) != incoming.get(field)
    ]


def _canonical_fingerprint(value: dict[str, Any]) -> str:
    encoded = json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _dedupe_snapshot(decisions: Sequence[_TradeDecision]) -> str:
    snapshot = [
        {
            "status": decision.status,
            "identity_kind": decision.identity_kind,
            "identity_value": decision.identity_value,
            "existing_event_ids": list(decision.existing_event_ids),
            "existing_fingerprints": list(decision.existing_fingerprints),
            "conflict": decision.conflict,
        }
        for decision in decisions
    ]
    encoded = json.dumps(snapshot, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _canonical_decimal(value: Any) -> str:
    numeric = Decimal(str(value)).quantize(_STORAGE_QUANT, rounding=ROUND_HALF_UP)
    return format(numeric, "f")


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        numeric = Decimal(str(value))
    except InvalidOperation:
        return None
    return numeric if numeric.is_finite() else None


def _timestamp_key(value: datetime) -> str:
    return _as_utc(value).isoformat(timespec="microseconds")


def _recover_concurrent_commit(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    preview_token: str,
    account_row_id: int | None,
) -> dict[str, Any] | None:
    try:
        account = _resolve_import_account(
            db,
            user_id=user_id,
            account_id=account_id,
            account_row_id=account_row_id,
        )
        staged = _get_staged_preview(
            db,
            token_hash=_token_hash(_validated_preview_token(preview_token)),
            user_id=user_id,
            account_id=account_id,
            account_row_id=int(account.id),
        )
        if staged is None:
            return None
        if staged.status == "committed" and staged.import_batch_id is not None:
            batch = _owned_batch_for_preview(db, staged)
            return _serialize_existing_batch(batch) if batch is not None else None
        existing_batch = (
            db.query(TradeImportBatch)
            .filter(TradeImportBatch.user_id == user_id)
            .filter(TradeImportBatch.account_id == account_id)
            .filter(TradeImportBatch.file_sha256 == staged.file_sha256)
            .one_or_none()
        )
        if existing_batch is None:
            return None
        _mark_preview_committed(staged, existing_batch, now=datetime.now(timezone.utc))
        db.commit()
        return _serialize_existing_batch(existing_batch)
    except Exception:
        db.rollback()
        return None


def _mark_preview_stale_after_integrity_conflict(
    db: Session,
    *,
    user_id: str,
    account_id: int,
    preview_token: str,
    account_row_id: int | None,
) -> None:
    try:
        account = _resolve_import_account(
            db,
            user_id=user_id,
            account_id=account_id,
            account_row_id=account_row_id,
        )
        staged = _get_staged_preview(
            db,
            token_hash=_token_hash(_validated_preview_token(preview_token)),
            user_id=user_id,
            account_id=account_id,
            account_row_id=int(account.id),
        )
        if staged is not None and staged.status != "committed":
            staged.status = "stale"
            staged.outcome_code = "preview_stale"
            staged.updated_at = datetime.now(timezone.utc)
            _scrub_staged_manifest(staged)
            db.commit()
    except Exception:
        db.rollback()


def _safe_failure_category(exc: Exception) -> str:
    if isinstance(exc, TradeImportValidationError):
        return f"validation:{exc.code}"
    if isinstance(exc, IntegrityError):
        return "integrity_error"
    return "internal_error"


def _elapsed_ms(started: float) -> float:
    return round(max((perf_counter() - started) * 1000.0, 0.0), 3)


def _log_import_outcome(event: str, **fields: Any) -> None:
    # Only a fixed metrics allowlist is logged. In particular, do not add file
    # names, hashes, tokens, trade ids, account/user ids, P&L, or row payloads.
    allowed = {
        "outcome",
        "total_rows",
        "new_rows",
        "duplicate_rows",
        "conflict_rows",
        "parse_ms",
        "dedupe_ms",
        "persist_ms",
        "commit_ms",
        "total_ms",
        "expired_previews",
        "deleted_previews",
        "failure_phase",
        "error_rows",
    }
    payload = {"event": f"trade_import_{event}"}
    payload.update(
        {
            key: fields[key]
            for key in sorted(fields)
            if key in allowed and fields[key] is not None
        }
    )
    logger.info("trade_import %s", json.dumps(payload, sort_keys=True, separators=(",", ":")))


def _parse_file(*, filename: str, content: bytes) -> _ParsedFile:
    source_file_name = _normalize_source_file_name(filename)
    content_bytes = _validate_content(content)
    extension = Path(source_file_name).suffix.lower()

    if extension in _CSV_EXTENSIONS:
        headers, rows = _read_csv_rows(content_bytes)
    elif extension in _EXCEL_EXTENSIONS:
        headers, rows = _read_excel_rows(content_bytes)
    else:
        raise TradeImportValidationError(
            "unsupported_file_type",
            "Unsupported trade file type. Upload a CSV or XLSX file.",
        )

    column_indexes = _resolve_columns(headers)
    parsed_trades: list[_ParsedTrade] = []
    row_errors: list[dict[str, Any]] = []

    for source_row in rows:
        if len(source_row.values) != len(headers):
            row_errors.append(
                {
                    "row_number": source_row.row_number,
                    "field": "row",
                    "message": (
                        f"Row {source_row.row_number} has {len(source_row.values)} values; "
                        f"expected {len(headers)}."
                    ),
                }
            )
            continue

        trade, errors = _parse_trade_row(
            source_row,
            headers=headers,
            column_indexes=column_indexes,
        )
        if errors:
            row_errors.extend(errors)
        elif trade is not None:
            parsed_trades.append(trade)

        if len(row_errors) >= MAX_REPORTED_ROW_ERRORS:
            break

    if row_errors:
        reported = row_errors[:MAX_REPORTED_ROW_ERRORS]
        first = reported[0]
        message = str(first.get("message") or "The trade file contains invalid rows.")
        if len(row_errors) > 1:
            message = f"{message} {len(row_errors) - 1} additional row error(s) were found."
        raise TradeImportValidationError(
            "invalid_rows",
            message,
            row_errors=reported,
        )

    if not parsed_trades:
        raise TradeImportValidationError(
            "no_trade_rows",
            "The trade file does not contain any trade rows.",
        )

    if len(parsed_trades) > MAX_TRADE_IMPORT_ROWS:
        raise TradeImportValidationError(
            "too_many_rows",
            f"The trade file exceeds the {MAX_TRADE_IMPORT_ROWS:,}-row import limit.",
        )

    return _ParsedFile(
        source_file_name=source_file_name,
        file_sha256=hashlib.sha256(content_bytes).hexdigest(),
        trades=tuple(parsed_trades),
    )


def _read_csv_rows(content: bytes) -> tuple[tuple[str, ...], tuple[_SourceRow, ...]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise TradeImportValidationError(
            "invalid_csv_encoding",
            "The CSV file must use UTF-8 encoding.",
        ) from exc

    try:
        dialect = _sniff_csv_dialect(text)
        reader = csv.reader(io.StringIO(text, newline=""), dialect=dialect)
        return _collect_tabular_rows(reader)
    except csv.Error as exc:
        raise TradeImportValidationError(
            "invalid_csv",
            "The CSV file could not be parsed.",
        ) from exc


def _sniff_csv_dialect(text: str) -> type[csv.Dialect] | csv.Dialect:
    sample = text[:16_384]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _read_excel_rows(content: bytes) -> tuple[tuple[str, ...], tuple[_SourceRow, ...]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles.numbers import is_datetime
    except ImportError as exc:
        raise TradeImportValidationError(
            "excel_support_unavailable",
            "Excel import support is unavailable on this server.",
        ) from exc

    _validate_excel_archive(content)

    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise TradeImportValidationError(
            "invalid_excel",
            "The XLSX file could not be parsed.",
        ) from exc

    try:
        worksheet = workbook.active
        return _collect_tabular_rows(
            (
                tuple(
                    _ExcelCellValue(
                        value=cell.value,
                        date_only=(
                            isinstance(cell.value, (date, datetime))
                            and is_datetime(str(cell.number_format or "")) == "date"
                        ),
                    )
                    for cell in row
                )
                for row in worksheet.iter_rows(values_only=False)
            )
        )
    finally:
        workbook.close()


def _validate_excel_archive(content: bytes) -> None:
    try:
        with ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            uncompressed_bytes = sum(entry.file_size for entry in entries)
    except (BadZipFile, OSError) as exc:
        raise TradeImportValidationError(
            "invalid_excel",
            "The XLSX file could not be parsed.",
        ) from exc

    if (
        len(entries) > MAX_EXCEL_ARCHIVE_ENTRIES
        or uncompressed_bytes > MAX_EXCEL_UNCOMPRESSED_BYTES
    ):
        raise TradeImportValidationError(
            "excel_archive_too_large",
            "The XLSX file expands beyond the safe import limit.",
        )


def _collect_tabular_rows(
    rows: Iterable[Sequence[Any]],
) -> tuple[tuple[str, ...], tuple[_SourceRow, ...]]:
    headers: tuple[str, ...] | None = None
    output: list[_SourceRow] = []

    for row_number, raw_row in enumerate(rows, start=1):
        wrapped_values = tuple(raw_row)
        date_only_indexes = frozenset(
            index
            for index, value in enumerate(wrapped_values)
            if isinstance(value, _ExcelCellValue) and value.date_only
        )
        values = tuple(
            value.value if isinstance(value, _ExcelCellValue) else value
            for value in wrapped_values
        )
        if _row_is_blank(values):
            continue

        if headers is None:
            headers = tuple(_header_text(value) for value in values)
            continue

        output.append(
            _SourceRow(
                row_number=row_number,
                values=values,
                date_only_indexes=date_only_indexes,
            )
        )
        if len(output) > MAX_TRADE_IMPORT_ROWS:
            raise TradeImportValidationError(
                "too_many_rows",
                f"The trade file exceeds the {MAX_TRADE_IMPORT_ROWS:,}-row import limit.",
            )

    if headers is None:
        raise TradeImportValidationError(
            "empty_file",
            "The trade file is empty.",
        )
    if not any(header.strip() for header in headers):
        raise TradeImportValidationError(
            "missing_header",
            "The trade file does not contain a header row.",
        )
    if any("\x00" in header for header in headers):
        raise TradeImportValidationError(
            "invalid_header",
            "Trade file headers cannot contain NUL characters.",
        )

    return headers, tuple(output)


def _resolve_columns(headers: Sequence[str]) -> dict[str, int]:
    matches: dict[str, int] = {}
    ambiguous: set[str] = set()

    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        for canonical, aliases in _COLUMN_ALIASES.items():
            if normalized not in aliases:
                continue
            if canonical in matches:
                ambiguous.add(canonical)
            else:
                matches[canonical] = index
            break

    if ambiguous:
        labels = [_COLUMN_LABELS[column] for column in sorted(ambiguous)]
        raise TradeImportValidationError(
            "ambiguous_columns",
            f"Multiple columns map to the same trade field: {', '.join(labels)}.",
        )

    missing = [
        _COLUMN_LABELS[column]
        for column in _REQUIRED_COLUMNS
        if column not in matches
    ]
    if missing:
        raise TradeImportValidationError(
            "missing_columns",
            f"Missing required columns: {', '.join(missing)}.",
            missing_columns=missing,
        )

    return matches


def _parse_trade_row(
    source_row: _SourceRow,
    *,
    headers: Sequence[str],
    column_indexes: dict[str, int],
) -> tuple[_ParsedTrade | None, list[dict[str, Any]]]:
    errors: list[dict[str, Any]] = []

    for index, value in enumerate(source_row.values):
        if isinstance(value, str) and "\x00" in value:
            field = headers[index] or f"Column {index + 1}"
            errors.append(
                {
                    "row_number": source_row.row_number,
                    "field": field,
                    "message": (
                        f"Row {source_row.row_number}, {field}: "
                        "cannot contain NUL characters."
                    ),
                }
            )
    if errors:
        return None, errors

    def raw(field: str) -> Any:
        index = column_indexes.get(field)
        return None if index is None else source_row.values[index]

    def capture(field: str, parser: Callable[[Any], Any]) -> Any:
        index = column_indexes.get(field)
        if field in {"entered_at", "exited_at"} and index in source_row.date_only_indexes:
            label = _COLUMN_LABELS[field]
            errors.append(
                {
                    "row_number": source_row.row_number,
                    "field": label,
                    "message": (
                        f"Row {source_row.row_number}, {label}: "
                        "must include a time; date-only spreadsheet cells are not supported."
                    ),
                }
            )
            return None
        try:
            return parser(raw(field))
        except ValueError as exc:
            label = _COLUMN_LABELS[field]
            errors.append(
                {
                    "row_number": source_row.row_number,
                    "field": label,
                    "message": f"Row {source_row.row_number}, {label}: {exc}",
                }
            )
            return None

    source_trade_id = capture("source_trade_id", _parse_identifier)
    contract_name = capture("contract_name", _parse_contract_name)
    entered_at = capture("entered_at", _parse_datetime_value)
    exited_at = capture("exited_at", _parse_datetime_value)
    entry_price = capture("entry_price", lambda value: _parse_decimal(value, positive=True))
    exit_price = capture("exit_price", lambda value: _parse_decimal(value, positive=True))
    fees = capture("fees", lambda value: _parse_decimal(value, nonnegative=True))
    gross_pnl = capture("pnl", _parse_decimal)
    size = capture("size", _parse_quantity)
    direction_result = capture("direction", _parse_direction)
    commissions = capture(
        "commissions",
        lambda value: _parse_decimal(value, nonnegative=True),
    )

    if isinstance(entered_at, datetime) and isinstance(exited_at, datetime) and entered_at > exited_at:
        errors.append(
            {
                "row_number": source_row.row_number,
                "field": _COLUMN_LABELS["exited_at"],
                "message": (
                    f"Row {source_row.row_number}, {_COLUMN_LABELS['exited_at']}: "
                    "must be on or after EnteredAt."
                ),
            }
        )

    trade_day = capture("trade_day", _parse_trade_day)
    if isinstance(exited_at, datetime) and isinstance(trade_day, date):
        derived_trade_day = trading_day_date(exited_at)
        if trade_day != derived_trade_day:
            errors.append(
                {
                    "row_number": source_row.row_number,
                    "field": _COLUMN_LABELS["trade_day"],
                    "message": (
                        f"Row {source_row.row_number}, {_COLUMN_LABELS['trade_day']}: "
                        f"must match the {derived_trade_day.isoformat()} futures trading day "
                        "derived from ExitedAt using the 6:00 PM ET boundary."
                    ),
                }
            )

    duration = _optional_display_text(raw("duration")) if "duration" in column_indexes else None
    if duration is None and isinstance(entered_at, datetime) and isinstance(exited_at, datetime):
        duration = _format_duration(exited_at - entered_at)

    if errors:
        return None, errors

    assert isinstance(source_trade_id, str)
    assert isinstance(contract_name, str)
    assert isinstance(entered_at, datetime)
    assert isinstance(exited_at, datetime)
    assert isinstance(entry_price, Decimal)
    assert isinstance(exit_price, Decimal)
    assert isinstance(fees, Decimal)
    assert isinstance(commissions, Decimal)
    assert isinstance(gross_pnl, Decimal)
    assert isinstance(size, Decimal)
    assert isinstance(direction_result, tuple)
    assert trade_day is not None

    direction, closing_side = direction_result
    symbol = normalize_symbol_key(contract_name) or contract_name.upper()
    raw_columns = {
        str(header): _json_safe_value(value)
        for header, value in zip(headers, source_row.values)
        if str(header).strip()
    }

    return (
        _ParsedTrade(
            row_number=source_row.row_number,
            source_trade_id=source_trade_id,
            contract_name=contract_name,
            symbol=symbol,
            entered_at=entered_at,
            exited_at=exited_at,
            entry_price=entry_price,
            exit_price=exit_price,
            fees=fees,
            commissions=commissions,
            gross_pnl=gross_pnl,
            size=size,
            direction=direction,
            closing_side=closing_side,
            trade_day=trade_day,
            duration=duration,
            raw_columns=raw_columns,
        ),
        [],
    )


def _parse_identifier(value: Any) -> str:
    if value is None or isinstance(value, bool):
        raise ValueError("is required.")
    if isinstance(value, int):
        text = str(value)
        return _validated_identifier(text)
    if isinstance(value, Decimal):
        if not value.is_finite() or value != value.to_integral_value():
            raise ValueError("must be a whole identifier.")
        return _validated_identifier(str(int(value)))
    if isinstance(value, float):
        if not math.isfinite(value) or not value.is_integer():
            raise ValueError("must be a whole identifier.")
        return _validated_identifier(str(int(value)))

    text = str(value).strip()
    if not text:
        raise ValueError("is required.")
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return _validated_identifier(text)


def _validated_identifier(value: str) -> str:
    if len(value) > MAX_SOURCE_TRADE_ID_CHARS:
        raise ValueError(
            f"must be {MAX_SOURCE_TRADE_ID_CHARS} characters or fewer."
        )
    if _has_control_character(value):
        raise ValueError("cannot contain control characters.")
    return value


def _parse_contract_name(value: Any) -> str:
    text = _required_text(value)
    if len(text) > 200:
        raise ValueError("must be 200 characters or fewer.")
    if _has_control_character(text):
        raise ValueError("cannot contain control characters.")
    return text.upper()


def _parse_decimal(
    value: Any,
    *,
    positive: bool = False,
    nonnegative: bool = False,
) -> Decimal:
    if value is None or isinstance(value, bool):
        raise ValueError("is required.")

    if isinstance(value, Decimal):
        numeric = value
    elif isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            raise ValueError("must be a finite number.")
        numeric = Decimal(str(value))
    else:
        text = str(value).strip()
        if not text:
            raise ValueError("is required.")
        is_parenthesized = text.startswith("(") and text.endswith(")")
        if is_parenthesized:
            text = text[1:-1]
        text = text.replace(",", "").replace("$", "").strip()
        try:
            numeric = Decimal(text)
        except InvalidOperation as exc:
            raise ValueError("must be a number.") from exc
        if is_parenthesized:
            numeric = -numeric

    if not numeric.is_finite():
        raise ValueError("must be a finite number.")

    try:
        normalized = numeric.quantize(_STORAGE_QUANT, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError("is outside the supported numeric range.") from exc
    if abs(normalized) > _STORAGE_MAX:
        raise ValueError("is outside the supported numeric range.")
    if positive and normalized <= 0:
        raise ValueError("must be greater than zero.")
    if nonnegative and normalized < 0:
        raise ValueError("must be zero or greater.")
    return normalized


def _parse_quantity(value: Any) -> Decimal:
    quantity = _parse_decimal(value, positive=True)
    if quantity != quantity.to_integral_value():
        raise ValueError("must be a whole number of futures contracts.")
    if quantity > MAX_FUTURES_QUANTITY:
        raise ValueError(f"must be {MAX_FUTURES_QUANTITY:,} contracts or fewer.")
    return quantity


def _parse_datetime_value(value: Any) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        raise ValueError("must include a time; date-only values are not supported.")
    else:
        text = _required_text(value)
        if re.fullmatch(
            r"(?:\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[/-]\d{1,2}[/-]\d{4})",
            text,
        ):
            raise ValueError("must include a time; date-only values are not supported.")
        iso_text = text[:-1] + "+00:00" if text.endswith(("Z", "z")) else text
        try:
            parsed = datetime.fromisoformat(iso_text)
        except ValueError:
            parsed = _parse_datetime_with_formats(text)

    if parsed.tzinfo is None:
        parsed = _localize_eastern_wall_time(parsed)
    return parsed.astimezone(timezone.utc)


def _localize_eastern_wall_time(value: datetime) -> datetime:
    naive = value.replace(tzinfo=None)
    fold_zero = naive.replace(tzinfo=_TRADING_TZ, fold=0)
    fold_one = naive.replace(tzinfo=_TRADING_TZ, fold=1)
    roundtrip_zero = fold_zero.astimezone(timezone.utc).astimezone(_TRADING_TZ).replace(tzinfo=None)
    roundtrip_one = fold_one.astimezone(timezone.utc).astimezone(_TRADING_TZ).replace(tzinfo=None)
    zero_valid = roundtrip_zero == naive
    one_valid = roundtrip_one == naive
    if not zero_valid and not one_valid:
        raise ValueError("is a nonexistent Eastern time during the spring DST transition.")
    if zero_valid and one_valid and fold_zero.utcoffset() != fold_one.utcoffset():
        raise ValueError(
            "is an ambiguous Eastern time during the fall DST transition; include an explicit UTC offset."
        )
    return fold_zero if zero_valid else fold_one


def _parse_datetime_with_formats(value: str) -> datetime:
    for date_format in _DATETIME_FORMATS:
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            continue
    raise ValueError("must be a valid date and time.")


def _parse_trade_day(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _required_text(value)
    leading_date = text.split(maxsplit=1)[0]
    for candidate in (leading_date, text):
        for date_format in _DATE_FORMATS:
            try:
                return datetime.strptime(candidate, date_format).date()
            except ValueError:
                continue

    try:
        parsed = _parse_datetime_value(text)
    except ValueError as exc:
        raise ValueError("must be a valid trade date.") from exc
    return parsed.astimezone(_TRADING_TZ).date()


def _parse_direction(value: Any) -> tuple[str, str]:
    normalized = re.sub(r"[^A-Z]", "", _required_text(value).upper())
    if normalized in {"LONG", "L", "BUY"}:
        # ProjectX closed-trade rows store the closing execution side.
        return "Long", "SELL"
    if normalized in {"SHORT", "S", "SELL"}:
        return "Short", "BUY"
    raise ValueError("must be Long or Short.")


def _summarize_trades(trades: Sequence[_ParsedTrade]) -> dict[str, Any]:
    gross_pnl = sum((trade.gross_pnl for trade in trades), Decimal("0"))
    fees = sum((trade.fees for trade in trades), Decimal("0"))
    commissions = sum((trade.commissions for trade in trades), Decimal("0"))
    net_values = [trade.net_pnl for trade in trades]
    return {
        "gross_pnl": _money_float(gross_pnl),
        "fees": _money_float(fees),
        "commissions": _money_float(commissions),
        "net_pnl": _money_float(gross_pnl - fees - commissions),
        "wins": sum(1 for value in net_values if value > 0),
        "losses": sum(1 for value in net_values if value < 0),
        "breakeven": sum(1 for value in net_values if value == 0),
    }


def _normalize_source_file_name(filename: str) -> str:
    text = str(filename or "").strip()
    source_file_name = re.split(r"[\\/]", text)[-1].strip()
    if not source_file_name:
        raise TradeImportValidationError(
            "missing_filename",
            "The trade file name is required.",
        )
    if len(source_file_name) > 255:
        raise TradeImportValidationError(
            "filename_too_long",
            "The trade file name must be 255 characters or fewer.",
        )
    if _has_control_character(source_file_name):
        raise TradeImportValidationError(
            "invalid_filename",
            "The trade file name cannot contain control characters.",
        )
    return source_file_name


def _validate_content(content: bytes) -> bytes:
    if not isinstance(content, (bytes, bytearray, memoryview)):
        raise TradeImportValidationError(
            "invalid_file_content",
            "The uploaded trade file is invalid.",
        )
    content_bytes = bytes(content)
    if not content_bytes:
        raise TradeImportValidationError(
            "empty_file",
            "The trade file is empty.",
        )
    if len(content_bytes) > MAX_TRADE_IMPORT_BYTES:
        raise TradeImportValidationError(
            "file_too_large",
            f"The trade file exceeds the {MAX_TRADE_IMPORT_BYTES // (1024 * 1024)} MB limit.",
        )
    return content_bytes


def _as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _normalize_header(value: Any) -> str:
    text = _header_text(value).lstrip("\ufeff").strip().lower()
    text = text.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", text)


def _header_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _required_text(value: Any) -> str:
    if value is None:
        raise ValueError("is required.")
    text = str(value).strip()
    if not text:
        raise ValueError("is required.")
    return text


def _has_control_character(value: str) -> bool:
    return any(ord(character) < 32 or ord(character) == 127 for character in value)


def _optional_display_text(value: Any) -> str | None:
    if _value_is_blank(value):
        return None
    if isinstance(value, timedelta):
        total_seconds = value.total_seconds()
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours):02d}:{int(minutes):02d}:{seconds:09.6f}"
    if isinstance(value, time):
        return value.isoformat()
    return str(value).strip()


def _format_duration(value: timedelta) -> str:
    total_microseconds = max(
        0,
        value.days * 86_400_000_000 + value.seconds * 1_000_000 + value.microseconds,
    )
    total_seconds, microseconds = divmod(total_microseconds, 1_000_000)
    hours, remainder = divmod(total_seconds, 3_600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{microseconds:06d}0"


def _row_is_blank(values: Sequence[Any]) -> bool:
    return all(_value_is_blank(value) for value in values)


def _value_is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    return str(value)


def _decimal_float(value: Decimal) -> float:
    return float(value)


def _money_float(value: Decimal) -> float:
    try:
        rounded = value.quantize(_MONEY_QUANT, rounding=ROUND_HALF_UP)
    except InvalidOperation:
        rounded = Decimal("0")
    return float(rounded)
